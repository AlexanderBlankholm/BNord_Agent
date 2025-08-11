import streamlit as st
import os
import json
from datetime import datetime
from simple_agent import SimpleAgent, ProjectData
from ai_agent import AIAgent

# Page configuration
st.set_page_config(
    page_title="Construction Project Agent",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .agent-card {
        border: 2px solid #e0e0e0;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        background-color: #f8f9fa;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .step-header {
        font-size: 1.3rem;
        font-weight: bold;
        color: #495057;
        margin: 1rem 0 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

def load_categories():
    """Load categories from JSON file"""
    try:
        with open("categories.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("categories", [])
    except FileNotFoundError:
        return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft"]

def create_excel_download(project_data):
    """Create Excel file and return as download button"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Projekt Budget"
        
        # Add headers
        headers = ["Kategori", "Opgave", "Beskrivelse"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add project info
        ws.cell(row=2, column=1, value="PROJEKT INFO")
        ws.cell(row=3, column=1, value="Beskrivelse")
        ws.cell(row=3, column=2, value=project_data.description)
        ws.cell(row=4, column=1, value="Kvadratmeter")
        ws.cell(row=4, column=2, value=project_data.square_meters)
        
        # Add tasks
        current_row = 6
        for category, tasks in project_data.category_tasks.items():
            for task in tasks:
                ws.cell(row=current_row, column=1, value=category)
                ws.cell(row=current_row, column=2, value=task)
                ws.cell(row=current_row, column=3, value=f"Opgave i {category}")
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"projekt_budget_{timestamp}.xlsx"
        
        return output, filename
        
    except Exception as e:
        st.error(f"Fejl ved oprettelse af Excel-fil: {e}")
        return None, None

def run_simple_agent_ui():
    """Streamlit UI for the simple agent"""
    st.markdown('<div class="main-header">🏗️ Simple Construction Agent</div>', unsafe_allow_html=True)
    
    # Initialize session state
    if 'simple_project_data' not in st.session_state:
        st.session_state.simple_project_data = ProjectData()
    
    # Step 1: Project Description
    st.markdown('<div class="step-header">📝 Trin 1: Projektbeskrivelse</div>', unsafe_allow_html=True)
    project_description = st.text_input(
        "Beskriv dit projekt",
        value=st.session_state.simple_project_data.description,
        placeholder="f.eks. 'Total renovation af badeværelse'"
    )
    
    if project_description:
        st.session_state.simple_project_data.description = project_description
        st.success(f"📋 Projekt: {project_description}")
    
    # Step 2: Square Meters
    st.markdown('<div class="step-header">❓ Trin 2: Yderligere oplysninger</div>', unsafe_allow_html=True)
    square_meters = st.number_input(
        "Hvor mange kvadratmeter er der?",
        min_value=0.1,
        max_value=1000.0,
        value=st.session_state.simple_project_data.square_meters or 8.5,
        step=0.1,
        format="%.1f"
    )
    
    if square_meters:
        st.session_state.simple_project_data.square_meters = square_meters
        st.success(f"📏 Kvadratmeter: {square_meters} m²")
    
    # Step 3: Category Selection
    st.markdown('<div class="step-header">🏗️ Trin 3: Vælg relevante kategorier</div>', unsafe_allow_html=True)
    categories = load_categories()
    
    # Display categories with checkboxes
    st.write("Vælg de relevante kategorier:")
    selected_categories = []
    
    # Create two columns for better layout
    col1, col2 = st.columns(2)
    for i, category in enumerate(categories):
        if i < len(categories) // 2:
            if col1.checkbox(f"{category}", key=f"cat_{i}"):
                selected_categories.append(category)
        else:
            if col2.checkbox(f"{category}", key=f"cat_{i}"):
                selected_categories.append(category)
    
    if selected_categories:
        st.session_state.simple_project_data.selected_categories = selected_categories
        st.success(f"✅ Valgte kategorier: {', '.join(selected_categories)}")
    
    # Step 4-7: Task Gathering
    if selected_categories:
        st.markdown('<div class="step-header">🔨 Trin 4-7: Opgaver for hver kategori</div>', unsafe_allow_html=True)
        
        for category in selected_categories:
            st.subheader(f"📋 Kategori: {category}")
            
            # Initialize tasks list for this category if not exists
            if category not in st.session_state.simple_project_data.category_tasks:
                st.session_state.simple_project_data.category_tasks[category] = []
            
            # Task input
            new_task = st.text_input(
                f"Hvilke underopgaver er nødvendige i {category}?",
                key=f"task_{category}",
                placeholder=f"f.eks. 'nedrivning af fliser på vægge'"
            )
            
            # Add task button
            if st.button(f"Tilføj opgave til {category}", key=f"add_{category}"):
                if new_task.strip():
                    st.session_state.simple_project_data.category_tasks[category].append(new_task.strip())
                    st.success(f"✅ Opgave tilføjet: {new_task}")
                    # Clear the input
                    st.rerun()
            
            # Display existing tasks
            if st.session_state.simple_project_data.category_tasks[category]:
                st.write("**Eksisterende opgaver:**")
                for i, task in enumerate(st.session_state.simple_project_data.category_tasks[category]):
                    col1, col2 = st.columns([4, 1])
                    col1.write(f"• {task}")
                    if col2.button("🗑️", key=f"del_{category}_{i}"):
                        st.session_state.simple_project_data.category_tasks[category].pop(i)
                        st.rerun()
    
    # Step 8: Excel Export
    if (st.session_state.simple_project_data.description and 
        st.session_state.simple_project_data.square_meters and 
        st.session_state.simple_project_data.selected_categories and
        any(st.session_state.simple_project_data.category_tasks.values())):
        
        st.markdown('<div class="step-header">💾 Trin 8: Gem til Excel</div>', unsafe_allow_html=True)
        
        # Project summary
        total_tasks = sum(len(tasks) for tasks in st.session_state.simple_project_data.category_tasks.values())
        st.info(f"""
        **Projekt sammendrag:**
        - Beskrivelse: {st.session_state.simple_project_data.description}
        - Kvadratmeter: {st.session_state.simple_project_data.square_meters} m²
        - Kategorier: {', '.join(st.session_state.simple_project_data.selected_categories)}
        - Total opgaver: {total_tasks}
        """)
        
        # Create and download Excel
        if st.button("📥 Download Excel-fil"):
            output, filename = create_excel_download(st.session_state.simple_project_data)
            if output:
                st.download_button(
                    label="💾 Download Excel",
                    data=output.read(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ Excel-fil klar til download!")

def run_ai_agent_ui():
    """Streamlit UI for the AI agent"""
    st.markdown('<div class="main-header">🤖 AI Construction Agent</div>', unsafe_allow_html=True)
    
    # Check OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        st.error("❌ OPENAI_API_KEY miljøvariabel er ikke sat.")
        st.info("Sæt venligst din OpenAI API nøgle før du bruger AI agenten.")
        return
    
    # Initialize session state
    if 'ai_project_data' not in st.session_state:
        st.session_state.ai_project_data = ProjectData()
    
    # Initialize AI agent
    if 'ai_agent' not in st.session_state:
        try:
            st.session_state.ai_agent = AIAgent()
            st.success("✅ AI Agent initialiseret!")
        except Exception as e:
            st.error(f"❌ Fejl ved initialisering af AI Agent: {e}")
            return
    
    # Step 1: Project Description
    st.markdown('<div class="step-header">📝 Trin 1: Projektbeskrivelse</div>', unsafe_allow_html=True)
    project_description = st.text_input(
        "Beskriv dit projekt",
        value=st.session_state.ai_project_data.description,
        placeholder="f.eks. 'Total renovation af badeværelse'",
        key="ai_project_desc"
    )
    
    if project_description:
        st.session_state.ai_project_data.description = project_description
        st.success(f"📋 Projekt: {project_description}")
        
        # AI follow-up questions
        if st.button("🤖 Få AI-forslag til opfølgende spørgsmål"):
            with st.spinner("AI genererer spørgsmål..."):
                try:
                    response = st.session_state.ai_agent.agent.run(
                        f"Brugeren har beskrevet deres projekt som: {project_description}. "
                        "Stil relevante opfølgende spørgsmål på dansk."
                    )
                    st.info(f"🤖 **AI-forslag:**\n\n{response}")
                except Exception as e:
                    st.error(f"Fejl ved AI-forespørgsel: {e}")
    
    # Step 2: Square Meters
    st.markdown('<div class="step-header">❓ Trin 2: Yderligere oplysninger</div>', unsafe_allow_html=True)
    square_meters = st.number_input(
        "Hvor mange kvadratmeter er der?",
        min_value=0.1,
        max_value=1000.0,
        value=st.session_state.ai_project_data.square_meters or 8.5,
        step=0.1,
        format="%.1f",
        key="ai_square_meters"
    )
    
    if square_meters:
        st.session_state.ai_project_data.square_meters = square_meters
        st.success(f"📏 Kvadratmeter: {square_meters} m²")
    
    # Step 3: Category Selection with AI assistance
    st.markdown('<div class="step-header">🏗️ Trin 3: Kategorier med AI-assistance</div>', unsafe_allow_html=True)
    categories = load_categories()
    
    # AI category recommendations
    if (st.session_state.ai_project_data.description and 
        st.session_state.ai_project_data.square_meters):
        
        if st.button("🤖 Få AI-anbefalinger til kategorier"):
            with st.spinner("AI analyserer dit projekt..."):
                try:
                    response = st.session_state.ai_agent.agent.run(
                        f"Præsenter disse kategorier for brugeren og forklar hvilke der kunne være "
                        f"relevante for et projekt om: {st.session_state.ai_project_data.description} "
                        f"på {st.session_state.ai_project_data.square_meters} m²"
                    )
                    st.info(f"🤖 **AI-anbefalinger:**\n\n{response}")
                except Exception as e:
                    st.error(f"Fejl ved AI-anbefaling: {e}")
    
    # Category selection
    st.write("Vælg de relevante kategorier:")
    selected_categories = []
    
    col1, col2 = st.columns(2)
    for i, category in enumerate(categories):
        if i < len(categories) // 2:
            if col1.checkbox(f"{category}", key=f"ai_cat_{i}"):
                selected_categories.append(category)
        else:
            if col2.checkbox(f"{category}", key=f"ai_cat_{i}"):
                selected_categories.append(category)
    
    if selected_categories:
        st.session_state.ai_project_data.selected_categories = selected_categories
        st.success(f"✅ Valgte kategorier: {', '.join(selected_categories)}")
    
    # Step 4-7: AI-guided task gathering
    if selected_categories:
        st.markdown('<div class="step-header">🔨 Trin 4-7: Opgaver med AI-assistance</div>', unsafe_allow_html=True)
        
        for category in selected_categories:
            st.subheader(f"📋 Kategori: {category}")
            
            # Initialize tasks list
            if category not in st.session_state.ai_project_data.category_tasks:
                st.session_state.ai_project_data.category_tasks[category] = []
            
            # AI task suggestions
            if st.button(f"🤖 Få AI-forslag til {category}", key=f"ai_suggest_{category}"):
                with st.spinner("AI genererer opgaveforslag..."):
                    try:
                        response = st.session_state.ai_agent.agent.run(
                            f"Spørg brugeren om hvilke specifikke opgaver der er nødvendige i "
                            f"kategorien '{category}' for et projekt om "
                            f"{st.session_state.ai_project_data.description}. "
                            f"Giv eksempler på typiske opgaver."
                        )
                        st.info(f"🤖 **AI-forslag for {category}:**\n\n{response}")
                    except Exception as e:
                        st.error(f"Fejl ved AI-forespørgsel: {e}")
            
            # Task input
            new_task = st.text_input(
                f"Hvilke underopgaver er nødvendige i {category}?",
                key=f"ai_task_{category}",
                placeholder=f"f.eks. 'nedrivning af fliser på vægge'"
            )
            
            # Add task button
            if st.button(f"Tilføj opgave til {category}", key=f"ai_add_{category}"):
                if new_task.strip():
                    st.session_state.ai_project_data.category_tasks[category].append(new_task.strip())
                    st.success(f"✅ Opgave tilføjet: {new_task}")
                    st.rerun()
            
            # Display existing tasks
            if st.session_state.ai_project_data.category_tasks[category]:
                st.write("**Eksisterende opgaver:**")
                for i, task in enumerate(st.session_state.ai_project_data.category_tasks[category]):
                    col1, col2 = st.columns([4, 1])
                    col1.write(f"• {task}")
                    if col2.button("🗑️", key=f"ai_del_{category}_{i}"):
                        st.session_state.ai_project_data.category_tasks[category].pop(i)
                        st.rerun()
    
    # Step 8: AI-generated summary and Excel export
    if (st.session_state.ai_project_data.description and 
        st.session_state.ai_project_data.square_meters and 
        st.session_state.ai_project_data.selected_categories and
        any(st.session_state.ai_project_data.category_tasks.values())):
        
        st.markdown('<div class="step-header">💾 Trin 8: AI-sammendrag og Excel</div>', unsafe_allow_html=True)
        
        # AI project summary
        if st.button("🤖 Generer AI-sammendrag"):
            with st.spinner("AI genererer projektsammendrag..."):
                try:
                    summary = f"""
                    Projekt: {st.session_state.ai_project_data.description}
                    Kvadratmeter: {st.session_state.ai_project_data.square_meters} m²
                    Kategorier: {', '.join(st.session_state.ai_project_data.selected_categories)}
                    Total opgaver: {sum(len(tasks) for tasks in st.session_state.ai_project_data.category_tasks.values())}
                    """.strip()
                    
                    response = st.session_state.ai_agent.agent.run(
                        f"Opret et detaljeret sammendrag af projektet på dansk. "
                        f"Projekt sammendrag: {summary}"
                    )
                    st.info(f"🤖 **AI-sammendrag:**\n\n{response}")
                except Exception as e:
                    st.error(f"Fejl ved AI-sammendrag: {e}")
        
        # Project summary display
        total_tasks = sum(len(tasks) for tasks in st.session_state.ai_project_data.category_tasks.values())
        st.info(f"""
        **Projekt sammendrag:**
        - Beskrivelse: {st.session_state.ai_project_data.description}
        - Kvadratmeter: {st.session_state.ai_project_data.square_meters} m²
        - Kategorier: {', '.join(st.session_state.ai_project_data.selected_categories)}
        - Total opgaver: {total_tasks}
        """)
        
        # Create and download Excel
        if st.button("📥 Download Excel-fil", key="ai_excel"):
            output, filename = create_excel_download(st.session_state.ai_project_data)
            if output:
                st.download_button(
                    label="💾 Download Excel",
                    data=output.read(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ Excel-fil klar til download!")

def main():
    """Main Streamlit app"""
    st.sidebar.title("🏗️ Construction Project Agent")
    st.sidebar.markdown("---")
    
    # Agent selection
    agent_choice = st.sidebar.radio(
        "Vælg agent:",
        ["Simple Agent", "AI Agent"],
        help="Simple Agent: Offline, ingen AI. AI Agent: LangChain + OpenAI integration."
    )
    
    # Sidebar info
    st.sidebar.markdown("---")
    st.sidebar.markdown("**📋 Funktioner:**")
    st.sidebar.markdown("• Projektbeskrivelse")
    st.sidebar.markdown("• Kategori-valg")
    st.sidebar.markdown("• Opgave-samling")
    st.sidebar.markdown("• Excel-eksport")
    
    if agent_choice == "Simple Agent":
        st.sidebar.markdown("**🔧 Simple Agent:**")
        st.sidebar.markdown("• Arbejder offline")
        st.sidebar.markdown("• Ingen API-omkostninger")
        st.sidebar.markdown("• Hurtig og pålidelig")
        run_simple_agent_ui()
    else:
        st.sidebar.markdown("**🤖 AI Agent:**")
        st.sidebar.markdown("• LangChain + OpenAI")
        st.sidebar.markdown("• Intelligente forslag")
        st.sidebar.markdown("• Kontekst-bevidst")
        st.sidebar.markdown("• Kræver API-nøgle")
        run_ai_agent_ui()
    
    # Footer
    st.sidebar.markdown("---")
    st.sidebar.markdown("**📚 Dokumentation:**")
    st.sidebar.markdown("[README](README.md)")
    st.sidebar.markdown("[Agent Summary](../AGENT_SUMMARY.md)")

if __name__ == "__main__":
    main()
