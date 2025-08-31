import streamlit as st
import os
import json
import time
from datetime import datetime
import sys

# Add the knowledge_base directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'embeddings_and_search'))

# Import the existing search system
from semantic_search import SemanticSearch

# Import the enhanced agent
from enhanced_ai_agent import EnhancedAIAgent, ProjectData

# Page configuration
st.set_page_config(
    page_title="Enhanced Construction Project Agent",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .search-result {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
        border-left: 4px solid #1f77b4;
    }
    .cost-breakdown {
        background-color: #e8f4fd;
        padding: 0.5rem;
        border-radius: 0.3rem;
        margin: 0.5rem 0;
    }
    .category-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #dee2e6;
    }
    .stButton > button {
        width: 100%;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

def load_categories():
    """Load categories from JSON file"""
    try:
        with open("categories.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("categories", [])
    except FileNotFoundError:
        st.warning("categories.json not found, using default categories")
        return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft", "døre", "vinduer", "belysning", "ventilation", "tætning", "projektledelse"]

def create_enhanced_excel_download(project_data: ProjectData):
    """Create enhanced Excel file with full cost breakdowns"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Projekt Budget"
        
        # Enhanced headers with cost structure
        headers = [
            "Kategori", "Opgave", "Beskrivelse", "Admin", "Kostpris_EP", "Materialer", 
            "Timer", "Takst", "Påslag_MAT", "Salgspris_MAT", "Påslag_UE", "Salgspris_UE", "Tilbud"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add project info
        ws.cell(row=2, column=1, value="PROJEKT INFO")
        ws.cell(row=3, column=1, value="Titel")
        ws.cell(row=3, column=2, value=project_data.description)
        ws.cell(row=4, column=1, value="Beskrivelse")
        ws.cell(row=4, column=2, value=getattr(project_data, 'detailed_description', ''))
        
        # Add components with full cost breakdown
        current_row = 6
        total_cost = 0
        
        for category, components in project_data.selected_components.items():
            for component in components:
                # Extract cost data from component
                ws.cell(row=current_row, column=1, value=component.get('kategori', category))
                ws.cell(row=current_row, column=2, value=component.get('Opgave', ''))
                ws.cell(row=current_row, column=3, value=f"Komponent fra {component.get('source_file', 'videnbase')}")
                ws.cell(row=current_row, column=4, value=component.get('Admin', 0))
                ws.cell(row=current_row, column=5, value=component.get('Kostpris_EP', 0))
                ws.cell(row=current_row, column=6, value=component.get('Materialer', 0))
                ws.cell(row=current_row, column=7, value=component.get('Timer', 0))
                ws.cell(row=current_row, column=8, value=component.get('Takst', 0))
                ws.cell(row=current_row, column=9, value=component.get('Påslag_MAT', 0))
                ws.cell(row=current_row, column=10, value=component.get('Salgspris_MAT', 0))
                ws.cell(row=current_row, column=11, value=component.get('Påslag_UE', 0))
                ws.cell(row=current_row, column=12, value=component.get('Salgspris_UE', 0))
                ws.cell(row=current_row, column=13, value=component.get('Tilbud', 0))
                
                total_cost += component.get('Tilbud', 0)
                current_row += 1
        
        # Add total row
        if current_row > 6:
            ws.cell(row=current_row, column=1, value="TOTAL")
            ws.cell(row=current_row, column=13, value=total_cost)
            
            # Style total row
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            total_font = Font(bold=True)
            for col in range(1, 14):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = total_fill
                cell.font = total_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes for download
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, f"enhanced_projekt_budget_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
    except Exception as e:
        st.error(f"Fejl ved oprettelse af Excel-fil: {e}")
        return None, None

def run_enhanced_agent_ui():
    """Enhanced Streamlit UI for the AI agent with knowledge base integration"""
    st.markdown('<h1 class="main-header">🤖 Enhanced AI Agent med RAG Pipeline</h1>', unsafe_allow_html=True)
    
    # Check OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        st.error("❌ OPENAI_API_KEY miljøvariabel er ikke sat. Sæt venligst din OpenAI API nøgle.")
        st.info("Du kan sætte den i PowerShell med: $env:OPENAI_API_KEY='din_nøgle_her'")
        return
    
    # Initialize session state with performance optimizations
    if 'project_data' not in st.session_state:
        st.session_state.project_data = ProjectData()
    
    # Cache expensive operations
    if 'search_system' not in st.session_state:
        with st.spinner("🚀 Initialiserer søgesystem (dette sker kun én gang)..."):
            try:
                # Initialize search system with trusted knowledge base path
                kb_path = os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'unified_knowledge_base_backup_before_bnord_fix.json')
                from component_embeddings import ComponentEmbeddings
                
                # Create embeddings manager with correct path and force regeneration
                embeddings_manager = ComponentEmbeddings(knowledge_base_path=kb_path)
                
                # Force regeneration of embeddings to ensure we have the corrected data
                with st.spinner("🔄 Regenererer embeddings med korrigeret data..."):
                    success = embeddings_manager.generate_embeddings(force_regenerate=True)
                    if success:
                        st.success("✅ Embeddings regenereret med korrigeret data")
                    else:
                        st.warning("⚠️ Kunne ikke regenerere embeddings, bruger cached data")
                
                # Create semantic search with the embeddings manager
                search_system = SemanticSearch(embeddings_manager=embeddings_manager)
                
                st.session_state.search_system = search_system
                st.success(f"✅ Søgesystem initialiseret med videnbase: {os.path.basename(kb_path)}")
            except Exception as e:
                st.error(f"❌ Kunne ikke initialisere søgesystem: {e}")
                st.session_state.search_system = None
    
    # Initialize AI agent (cache this too)
    if 'ai_agent' not in st.session_state:
        with st.spinner("🤖 Initialiserer AI Agent (dette sker kun én gang)..."):
            try:
                ai_agent = EnhancedAIAgent()
                st.session_state.ai_agent = ai_agent
                st.success("✅ AI Agent initialiseret med RAG-funktionalitet")
            except Exception as e:
                st.error(f"❌ Kunne ikke initialisere AI Agent: {e}")
                st.session_state.ai_agent = None
    
    # Step 0: Project Title
    st.header("📝 Trin 0: Titel")
    
    project_title = st.text_input(
        "Projekt titel",
        value=st.session_state.project_data.description,
        placeholder="F.eks. 'Total renovation af badeværelse'",
        help="Dette bliver titlen på den genererede Excel-fil"
    )
    
    if project_title:
        st.session_state.project_data.description = project_title
        st.success(f"✅ Projekt titel: {project_title}")
    
    # Step 1: Project Description
    if project_title:
        st.header("📋 Trin 1: Projektbeskrivelse")
        
        project_description = st.text_area(
            "Beskriv dit projekt i detaljer",
            value=getattr(st.session_state.project_data, 'detailed_description', ''),
            placeholder="F.eks. 'Total renovation af badeværelse med nye fliser og VVS. Projektet omfatter nedrivning af eksisterende fliser, installation af nye VVS-fittings, og lægning af nye fliser på både vægge og gulv.'",
            height=100,
            help="Dette kan bruges til at forbedre AI-generering af komponenter"
        )
        
        if project_description:
            st.session_state.project_data.detailed_description = project_description
            st.info(f"📝 Projektbeskrivelse gemt ({len(project_description)} tegn)")
    
    # Step 2: Category Selection
    if st.session_state.project_data.description:
        st.header("🏗️ Trin 2: Vælg Relevante Kategorier")
        
        categories = load_categories()
        
        # Display categories with descriptions
        st.info("Vælg de kategorier der er relevante for dit projekt:")
        
        selected_categories = []
        cols = st.columns(3)
        
        for i, category in enumerate(categories):
            col_idx = i % 3
            with cols[col_idx]:
                if st.checkbox(f"**{category.title()}**", key=f"cat_{category}"):
                    selected_categories.append(category)
        
        st.session_state.project_data.selected_categories = selected_categories
        
        if selected_categories:
            st.success(f"✅ Valgt {len(selected_categories)} kategorier: {', '.join(selected_categories)}")
    
    # Step 3: RAG-Based Component Generation
    if st.session_state.project_data.selected_categories:
        st.header("🤖 Trin 3: AI-Genererede Komponenter med RAG Pipeline")
        
        # Initialize components if not exists
        if 'selected_components' not in st.session_state.project_data.__dict__:
            st.session_state.project_data.selected_components = {}
        
        # Add a loading state to prevent unnecessary re-renders
        if 'rag_loading' not in st.session_state:
            st.session_state.rag_loading = {}
        
        for category in st.session_state.project_data.selected_categories:
            st.markdown(f'<div class="category-section">', unsafe_allow_html=True)
            st.subheader(f"📋 {category.title()}")
            
            # Initialize components list for this category
            if category not in st.session_state.project_data.selected_components:
                st.session_state.project_data.selected_components[category] = []
            
            # Task input for RAG generation
            task_input = st.text_input(
                f"Beskriv opgaven i {category} for AI-generering",
                key=f"task_{category}",
                placeholder=f"F.eks. 'nedrivning af fliser på vægge' for {category}"
            )
            
            # Generate button - only show if task input is provided
            if task_input:
                col1, col2 = st.columns([1, 3])
                with col1:
                    generate_button = st.button("🤖 Generer Komponent", key=f"generate_{category}")
                
                # Only proceed with generation if button is clicked
                if generate_button:
                    # Check if search system is available
                    if not st.session_state.search_system:
                        st.error("⚠️ Søgesystemet er ikke tilgængeligt. Prøv at genindlæse siden.")
                        continue
                    
                    # Check if AI agent is available
                    if not hasattr(st.session_state, 'ai_agent') or not st.session_state.ai_agent:
                        st.error("⚠️ AI Agent er ikke tilgængelig. Prøv at genindlæse siden.")
                        continue
                    
                    try:
                        # Generate AI component using RAG with detailed progress
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # Step 1: Search
                        status_text.text("🔍 Søger efter relevante komponenter...")
                        progress_bar.progress(25)
                        
                        # Step 2: Generate
                        status_text.text("🤖 Genererer AI-komponent med RAG...")
                        progress_bar.progress(50)
                        
                        # Generate the component
                        generated_component = st.session_state.ai_agent.generate_component_with_rag(
                            task_input, 
                            use_high_quality_only=st.session_state.use_high_quality_only
                        )
                        
                        # Store the generated component in session state to persist it
                        st.session_state[f"generated_component_{category}"] = generated_component
                        
                        # Step 3: Complete
                        progress_bar.progress(100)
                        status_text.text("✅ AI-komponent genereret!")
                        
                        # Clear progress indicators
                        progress_bar.empty()
                        status_text.empty()
                        
                        st.success("✅ AI-komponent genereret!")
                        
                    except Exception as e:
                        st.error(f"Fejl ved AI-generering: {e}")
                        st.info("💡 Prøv at genindlæse siden eller kontakt support")
                
                # Display component if we have one (either newly generated or stored)
                current_component = st.session_state.get(f"generated_component_{category}")
                if current_component and 'error' not in current_component:
                    generated_component = current_component
                    
                    # Display the generated component for review
                    st.subheader("🤖 Genereret AI-Komponent")
                    
                    # Component details - Show ALL fields
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        st.markdown(f"""
                        **Opgave:** {generated_component.get('Opgave', 'N/A')}
                        **Kategori:** {generated_component.get('kategori', 'N/A')}
                        **Fag:** {generated_component.get('Fag', 'N/A')}
                        **Kilde:** {generated_component.get('source_file', 'AI-genereret')}
                        """)
                    
                    with col2:
                        st.metric("Total Pris", f"{generated_component.get('Tilbud', 0):,.0f} DKK")
                    
                    # Detailed cost breakdown - Show ALL fields
                    st.markdown("**📊 Detaljeret Omkostningsstruktur:**")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Admin", f"{generated_component.get('Admin', 0):,.0f} DKK")
                        st.metric("Kostpris EP", f"{generated_component.get('Kostpris_EP', 0):,.0f} DKK")
                        st.metric("Materialer", f"{generated_component.get('Materialer', 0):,.0f} DKK")
                    with col2:
                        st.metric("Timer", f"{generated_component.get('Timer', 0):.1f}")
                        st.metric("Takst", f"{generated_component.get('Takst', 0):,.0f} DKK")
                        st.metric("Påslag MAT", f"{generated_component.get('Påslag_MAT', 0):.1f}%")
                    with col3:
                        st.metric("UE", f"{generated_component.get('UE', 0):,.0f} DKK")
                        st.metric("Påslag UE", f"{generated_component.get('Påslag_UE', 0):.1f}%")
                        st.metric("Salgspris MAT", f"{generated_component.get('Salgspris_MAT', 0):,.0f} DKK")
                    with col4:
                        st.metric("Salgspris UE", f"{generated_component.get('Salgspris_UE', 0):,.0f} DKK")
                        st.metric("Tilbud", f"{generated_component.get('Tilbud', 0):,.0f} DKK")
                    
                    # Show LLM generation status
                    st.success("🤖 **AI-Generated Component**: This component was created entirely by the LLM based on context components")
                    
                    # Action buttons for generated component
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        if st.button("✅ Tilføj Genereret Komponent", key=f"add_ai_{category}_{task_input}"):
                            # Ensure the category exists in selected_components
                            if category not in st.session_state.project_data.selected_components:
                                st.session_state.project_data.selected_components[category] = []
                            
                            # Add the component
                            st.session_state.project_data.selected_components[category].append(generated_component)
                            
                            st.success(f"✅ Tilføjet AI-komponent: {generated_component.get('Opgave', 'N/A')}")
                            
                            # Use session state to track that we've added a component
                            st.session_state[f"component_added_{category}"] = True
                    
                    with col2:
                        if st.button("🔄 Generer Ny Komponent", key=f"regenerate_{category}_{task_input}"):
                            # Clear the generated component to force regeneration
                            if f"generated_component_{category}" in st.session_state:
                                del st.session_state[f"generated_component_{category}"]
                    
                    with col3:
                        if st.button("🗑️ Ryd Cache", key=f"clear_cache_{category}_{task_input}"):
                            # Clear any cached data for this category
                            for key in list(st.session_state.keys()):
                                if key.startswith(f"manual_search_{category}"):
                                    del st.session_state[key]
                            st.success("✅ Cache ryddet for denne kategori")
                    
                    # Show context quality information
                    context_quality = generated_component.get('context_quality', 'unknown')
                    context_source = generated_component.get('context_source', 'unknown')
                    
                    if context_quality == 'high':
                        st.success(f"✅ **Høj Kvalitet**: {context_source}")
                    elif context_quality == 'mixed':
                        st.warning(f"⚠️ **Blandet Kvalitet**: {context_source}")
                    else:
                        st.info(f"ℹ️ **Kvalitet**: {context_source}")
                    
                    # Show context components used for generation (lazy loaded for performance)
                    with st.expander("🔍 Se kontekst-komponenter brugt til generering"):
                        # Use the actual context components from the RAG generation
                        context_components = generated_component.get('context_components_used', [])
                        
                        if context_components:
                            # Show summary first for quick loading
                            st.info(f"AI-komponenten er baseret på {len(context_components)} lignende komponenter fra videnbase:")
                            
                            # Lazy load detailed context (only when expanded)
                            show_details = st.checkbox("📋 Vis detaljerede kontekst-komponenter", key=f"show_details_{category}_{task_input}")
                            
                            if show_details:
                                for i, result in enumerate(context_components):
                                    quality_score = result.get('quality_score', 0.5)
                                    source_file = result.get('source_file', 'N/A')
                                    
                                    if quality_score == 1.0:
                                        quality_indicator = "🟢 Høj"
                                        quality_description = f"Fra {source_file}"
                                    else:
                                        quality_indicator = "🟡 Medium"
                                        quality_description = f"Transformed data fra {source_file}"
                                    
                                    # Show full field breakdown for each context component
                                    with st.expander(f"📊 Detaljer for {result.get('Opgave', 'N/A')[:30]}...", expanded=False):
                                        col1, col2, col3, col4 = st.columns(4)
                                        with col1:
                                            st.metric("Admin", f"{result.get('Admin', 0):,.0f} DKK")
                                            st.metric("Kostpris EP", f"{result.get('Kostpris_EP', 0):,.0f} DKK")
                                            st.metric("Materialer", f"{result.get('Materialer', 0):,.0f} DKK")
                                        with col2:
                                            st.metric("Timer", f"{result.get('Timer', 0):.1f}")
                                            st.metric("Takst", f"{result.get('Takst', 0):,.0f} DKK")
                                            st.metric("Påslag MAT", f"{result.get('Påslag_MAT', 0):.1f}%")
                                        with col3:
                                            st.metric("UE", f"{result.get('UE', 0):,.0f} DKK")
                                            st.metric("Påslag UE", f"{result.get('Påslag_UE', 0):.1f}%")
                                            st.metric("Salgspris MAT", f"{result.get('Salgspris_MAT', 0):,.0f} DKK")
                                        with col4:
                                            st.metric("Salgspris UE", f"{result.get('Salgspris_UE', 0):,.0f} DKK")
                                            st.metric("Tilbud", f"{result.get('Tilbud', 0):,.0f} DKK")
                                    
                                    # Summary line
                                    st.markdown(f"""
                                    **{i+1}. {result.get('Opgave', 'N/A')}**
                                    - Kategori: {result.get('kategori', 'N/A')}
                                    - Fag: {result.get('Fag', 'N/A')}
                                    - Pris: {result.get('Tilbud', 0):,.0f} DKK
                                    - Lighed: {result.get('similarity_score', 0):.3f}
                                    - Kvalitet: {quality_indicator}
                                    - Kilde: {quality_description}
                                    """)
                            else:
                                st.info("✅ Klik checkbox for at se detaljerede kontekst-komponenter")
                        else:
                            st.warning("Ingen kontekst-komponenter fundet")
            
            # Show success message if component was just added
            if st.session_state.get(f"component_added_{category}", False):
                st.success(f"✅ Komponent tilføjet til {category}!")
                # Clear the flag
                st.session_state[f"component_added_{category}"] = False
            
            # Show selected components for this category
            if st.session_state.project_data.selected_components.get(category):
                st.subheader(f"✅ Valgte komponenter for {category}")
                for i, component in enumerate(st.session_state.project_data.selected_components[category]):
                    col1, col2, col3 = st.columns([3, 2, 1])
                    with col1:
                        st.write(f"**{component.get('Opgave', 'N/A')}**")
                    with col2:
                        st.metric("Pris", f"{component.get('Tilbud', 0):,.0f} DKK")
                    with col3:
                        if st.button("🗑️", key=f"remove_{category}_{i}"):
                            st.session_state.project_data.selected_components[category].pop(i)
                            st.success(f"✅ Fjernet komponent: {component.get('Opgave', 'N/A')}")
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Step 4: "Anything else?" - Manual Component Selection
    if st.session_state.project_data.selected_categories:
        st.header("🔍 Trin 4: Er der andet du vil tilføje?")
        st.info("💡 Her kan du manuelt vælge komponenter fra videnbasen hvis AI-genereringen ikke dækkede alt du havde brug for.")
        
        # Manual search and selection with caching
        manual_search = st.text_input(
            "Søg manuelt efter komponenter",
            placeholder="F.eks. 'fliser' eller 'vvs'",
            key="manual_search"
        )
        
        if manual_search:
            if st.session_state.search_system:
                try:
                    # Cache search results to avoid re-searching
                    cache_key = f"manual_search_{manual_search}"
                    if cache_key not in st.session_state:
                        with st.spinner(f"🔍 Søger efter '{manual_search}'..."):
                            manual_results = st.session_state.search_system.search(manual_search, top_k=5, min_similarity=0.2)
                            st.session_state[cache_key] = manual_results
                    else:
                        manual_results = st.session_state[cache_key]
                        st.info(f"📋 Viser cached søgeresultater for '{manual_search}'")
                    
                    if manual_results:
                        st.success(f"Fandt {len(manual_results)} komponenter for '{manual_search}'")
                        
                        # Display manual search results
                        for i, result in enumerate(manual_results):
                            with st.expander(f"🔍 {result.get('Opgave', 'N/A')[:60]}..."):
                                st.markdown(f"""
                                **Opgave:** {result.get('Opgave', 'N/A')}
                                **Kategori:** {result.get('kategori', 'N/A')}
                                **Lighedsscore:** {result.get('similarity_score', 0):.3f}
                                """)
                                
                                # Cost breakdown
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Total Pris", f"{result.get('Tilbud', 0):,.0f} DKK")
                                    st.metric("Kostpris EP", f"{result.get('Kostpris_EP', 0):,.0f} DKK")
                                with col2:
                                    st.metric("Materialer", f"{result.get('Materialer', 0):,.0f} DKK")
                                    st.metric("Timer", f"{result.get('Timer', 0):.1f}")
                                with col3:
                                    st.metric("Takst", f"{result.get('Takst', 0):,.0f} DKK")
                                    st.metric("Påslag MAT", f"{result.get('Påslag_MAT', 0):.0f} DKK")
                                
                                # Category selection for manual components
                                selected_category = st.selectbox(
                                    "Vælg kategori for denne komponent:",
                                    st.session_state.project_data.selected_categories,
                                    key=f"cat_select_{i}"
                                )
                                
                                # Add to project button
                                if st.button(f"Tilføj til {selected_category}", key=f"add_manual_{i}"):
                                    if selected_category not in st.session_state.project_data.selected_components:
                                        st.session_state.project_data.selected_components[selected_category] = []
                                    st.session_state.project_data.selected_components[selected_category].append(result)
                                    st.success(f"✅ Tilføjet til {selected_category}: {result.get('Opgave', 'N/A')}")
                                    
                                    # Use session state to track that we've added a component
                                    st.session_state[f"component_added_{selected_category}"] = True
                        
                    else:
                        st.warning("Ingen komponenter fundet for din søgning. Prøv andre søgeord.")
                
                except Exception as e:
                    st.error(f"Fejl ved manuel søgning: {e}")
            else:
                st.error("⚠️ Søgesystemet er ikke tilgængeligt")
    
    # Step 5: Project Summary and Excel Export
    if st.session_state.project_data.selected_categories:
        st.header("💾 Trin 5: Projekt Sammendrag og Excel Export")
        
        # Calculate totals
        total_components = sum(len(components) for components in st.session_state.project_data.selected_components.values())
        total_cost = sum(
            sum(comp.get('Tilbud', 0) for comp in components)
            for components in st.session_state.project_data.selected_components.values()
        )
        
        # Display summary
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Kategorier", len(st.session_state.project_data.selected_categories))
            st.metric("Total Komponenter", total_components)
        with col2:
            st.metric("Total Omkostning", f"{total_cost:,.0f} DKK")
            st.metric("Gns. per Komponent", f"{total_cost/total_components:,.0f} DKK" if total_components > 0 else "0 DKK")
        with col3:
            st.metric("Projekt Titel", st.session_state.project_data.description[:30] + "..." if len(st.session_state.project_data.description) > 30 else st.session_state.project_data.description)
            st.metric("Beskrivelse", f"{len(getattr(st.session_state.project_data, 'detailed_description', ''))} tegn")
        
        # Excel export
        if total_components > 0:
            st.subheader("📊 Excel Export med Fuld Omkostningsstruktur")
            
            excel_data, filename = create_enhanced_excel_download(st.session_state.project_data)
            
            if excel_data:
                st.download_button(
                    label="📥 Download Enhanced Excel Budget",
                    data=excel_data.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download Excel-fil med fuld omkostningsstruktur inkl. Kostpris_EP, Materialer, Timer, Takst, Påslag og Salgspris felter"
                )
                
                st.success(f"✅ Excel-fil klar til download: {filename}")
                st.info("📋 Filen indeholder alle omkostningsdetaljer fra videnbasen med professionel formatering.")
            else:
                st.error("❌ Kunne ikke oprette Excel-fil")
        else:
            st.warning("⚠️ Ingen komponenter valgt endnu. Tilføj komponenter fra søgeresultaterne for at kunne eksportere til Excel.")

def main():
    """Main function"""
    # Initialize session state for quality control at the top level
    if 'use_high_quality_only' not in st.session_state:
        st.session_state.use_high_quality_only = True  # Default to high quality
    
    st.sidebar.title("🏗️ Enhanced Construction Agent")
    
    # Sidebar navigation
    st.sidebar.markdown("### 🤖 Enhanced AI Agent")
    st.sidebar.info("RAG-powered construction project planning with AI-generated components")
    
    # Quality control toggle
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🎯 Quality Control")
    
    use_high_quality_only = st.sidebar.checkbox(
        "✅ High-Quality Only (Default)",
        value=st.session_state.use_high_quality_only,
        help="When enabled, RAG generation only uses high-quality components. When disabled, uses entire database including transformed data."
    )
    
    # Update session state
    st.session_state.use_high_quality_only = use_high_quality_only
    
    # Quality status indicator
    if use_high_quality_only:
        st.sidebar.success("🔒 High-Quality Mode: Only reliable pricing data used")
    else:
        st.sidebar.warning("⚠️ Full Database Mode: All components available (quality may vary)")
    
    # Sidebar info
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🔍 RAG Pipeline Features")
    st.sidebar.markdown("""
    - **AI-Generated Components** baseret på videnbase
    - **Semantic Search** gennem 633 komponenter
    - **Smart Pricing** arver fra lignende komponenter
    - **Manual Backup** søgning tilgængelig
    """)
    
    st.sidebar.markdown("### 📊 Omkostningsfelter")
    st.sidebar.markdown("""
    - Kostpris_EP
    - Materialer
    - Timer & Takst
    - Påslag_MAT/UE
    - Salgspris_MAT/UE
    - Total Tilbud
    """)
    
    # Performance management
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚡ Performance")
    
    # Cache status
    cache_keys = [k for k in st.session_state.keys() if k.startswith('manual_search_')]
    st.sidebar.info(f"📋 Cached searches: {len(cache_keys)}")
    
    # Clear all cache button
    if st.sidebar.button("🗑️ Clear All Cache", help="Clear all cached search results to free memory"):
        for key in cache_keys:
            del st.session_state[key]
        st.sidebar.success("✅ All cache cleared!")
        st.rerun()
    
    # Performance tips
    st.sidebar.markdown("""
    **💡 Performance Tips:**
    - First run is slow (initialization)
    - Subsequent searches are cached
    - Clear cache if memory usage is high
    """)
    
    # Performance monitoring
    if 'performance_start_time' not in st.session_state:
        st.session_state.performance_start_time = None
    
    if st.sidebar.button("📊 Performance Info", help="Show performance statistics"):
        if st.session_state.performance_start_time:
            elapsed = time.time() - st.session_state.performance_start_time
            st.sidebar.metric("Session Time", f"{elapsed:.1f}s")
        else:
            st.session_state.performance_start_time = time.time()
            st.sidebar.info("Performance monitoring started")
    
    # Check OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        st.sidebar.error("❌ OPENAI_API_KEY mangler")
        st.sidebar.info("Sæt miljøvariabel for at bruge AI-funktioner")
    
    # Page routing
    run_enhanced_agent_ui()

if __name__ == "__main__":
    main()
