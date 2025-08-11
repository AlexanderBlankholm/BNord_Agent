import os
import json
from typing import Dict, List, Any
from dataclasses import dataclass
from datetime import datetime

@dataclass
class ProjectData:
    """Data structure to hold project information"""
    description: str = ""
    square_meters: float = 0.0
    selected_categories: List[str] = None
    category_tasks: Dict[str, List[str]] = None
    
    def __post_init__(self):
        if self.selected_categories is None:
            self.selected_categories = []
        if self.category_tasks is None:
            self.category_tasks = {}

class SimpleAgent:
    """A simple agent for gathering construction project information"""
    
    def __init__(self):
        self.project_data = ProjectData()
        self.categories = self.load_categories()
        
    def load_categories(self) -> List[str]:
        """Load categories from JSON file"""
        try:
            with open("categories.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("categories", [])
        except FileNotFoundError:
            print("Warning: categories.json not found, using default categories")
            return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft"]
    
    def run(self):
        """Main agent flow"""
        print("🤖 Hej! Jeg er din assistent til at oprette et budget for dit byggeprojekt.")
        print("=" * 60)
        
        # Step 1: Get project description
        self.get_project_description()
        
        # Step 2: Ask follow-up questions
        self.ask_follow_up_questions()
        
        # Step 3: Select relevant categories
        self.select_categories()
        
        # Step 4-7: Get tasks for each category
        self.get_category_tasks()
        
        # Step 8: Save to Excel
        self.save_to_excel()
        
        print("\n✅ Projektet er nu færdigt! Excel-filen er oprettet.")
    
    def get_project_description(self):
        """Step 1: Get initial project description"""
        print("\n📝 **Trin 1: Projektbeskrivelse**")
        self.project_data.description = input("Beskriv dit projekt (f.eks. 'Total renovation af badeværelse'): ")
        print(f"📋 Projekt: {self.project_data.description}")
    
    def ask_follow_up_questions(self):
        """Step 2: Ask follow-up questions"""
        print("\n❓ **Trin 2: Yderligere oplysninger**")
        
        # Ask for square meters
        while True:
            try:
                square_meters_input = input("Hvor mange kvadratmeter er der? (f.eks. 8.5): ")
                self.project_data.square_meters = float(square_meters_input)
                break
            except ValueError:
                print("❌ Indtast venligst et gyldigt tal (f.eks. 8.5)")
        
        print(f"📏 Kvadratmeter: {self.project_data.square_meters} m²")
    
    def select_categories(self):
        """Step 3: Select relevant categories"""
        print("\n🏗️ **Trin 3: Vælg relevante kategorier**")
        print("Her er de tilgængelige kategorier:")
        
        for i, category in enumerate(self.categories, 1):
            print(f"  {i}. {category}")
        
        print("\nIndtast numrene på de relevante kategorier (adskilt med komma, f.eks. 1,3,5):")
        
        while True:
            try:
                selection = input("Valgte kategorier: ")
                selected_indices = [int(x.strip()) - 1 for x in selection.split(",")]
                
                # Validate indices
                if all(0 <= idx < len(self.categories) for idx in selected_indices):
                    self.project_data.selected_categories = [self.categories[idx] for idx in selected_indices]
                    break
                else:
                    print("❌ Ugyldige numre. Prøv igen.")
            except ValueError:
                print("❌ Indtast venligst gyldige numre adskilt med komma.")
        
        print(f"✅ Valgte kategorier: {', '.join(self.project_data.selected_categories)}")
    
    def get_category_tasks(self):
        """Step 4-7: Get tasks for each selected category"""
        print("\n🔨 **Trin 4-7: Opgaver for hver kategori**")
        
        for category in self.project_data.selected_categories:
            print(f"\n📋 **Kategori: {category}**")
            tasks = []
            
            while True:
                task = input(f"Hvilke underopgaver er nødvendige i {category}? (f.eks. 'nedrivning af fliser på vægge'): ")
                if task.strip():
                    tasks.append(task.strip())
                
                more = input("Er der andet du vil tilføje til denne kategori? (ja/nej): ").lower().strip()
                if more not in ['ja', 'j', 'yes', 'y']:
                    break
            
            self.project_data.category_tasks[category] = tasks
            print(f"✅ {len(tasks)} opgaver tilføjet til {category}")
    
    def save_to_excel(self):
        """Step 8: Save project data to Excel"""
        print("\n💾 **Trin 8: Gem til Excel**")
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            # Create a simple Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Projekt Budget"
            
            # Add headers
            headers = ["Kategori", "Opgave", "Beskrivelse"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add project info
            ws.cell(row=2, column=1, value="PROJEKT INFO")
            ws.cell(row=3, column=1, value="Beskrivelse")
            ws.cell(row=3, column=2, value=self.project_data.description)
            ws.cell(row=4, column=1, value="Kvadratmeter")
            ws.cell(row=4, column=2, value=self.project_data.square_meters)
            
            # Add tasks
            current_row = 6
            for category, tasks in self.project_data.category_tasks.items():
                for task in tasks:
                    ws.cell(row=current_row, column=1, value=category)
                    ws.cell(row=current_row, column=2, value=task)
                    ws.cell(row=current_row, column=3, value=f"Opgave i {category}")
                    current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"projekt_budget_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"✅ Excel-fil gemt som: {filename}")
            
        except ImportError as e:
            print(f"❌ Fejl: Manglende bibliotek - {e}")
            print("Installer venligst: pip install pandas openpyxl")
        except Exception as e:
            print(f"❌ Fejl ved oprettelse af Excel-fil: {e}")

def main():
    """Main function to run the agent"""
    agent = SimpleAgent()
    agent.run()

if __name__ == "__main__":
    main()
