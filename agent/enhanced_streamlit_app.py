import streamlit as st
import os
import json
from datetime import datetime
import sys

# Add the knowledge_base directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'embeddings_and_search'))

# Import the existing search system
from semantic_search import SemanticSearch

# Import the enhanced agent
from enhanced_ai_agent import EnhancedAIAgent, ProjectData

# Page configuration
st.set_page_config(
    page_title="Enhanced Construction Project Agent",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .search-result {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
        border-left: 4px solid #1f77b4;
    }
    .cost-breakdown {
        background-color: #e8f4fd;
        padding: 0.5rem;
        border-radius: 0.3rem;
        margin: 0.5rem 0;
    }
    .category-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #dee2e6;
    }
    .stButton > button {
        width: 100%;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

def load_categories():
    """Load categories from JSON file"""
    try:
        with open("categories.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("categories", [])
    except FileNotFoundError:
        st.warning("categories.json not found, using default categories")
        return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft", "døre", "vinduer", "belysning", "ventilation", "tætning", "projektledelse"]

def create_enhanced_excel_download(project_data: ProjectData):
    """Create enhanced Excel file with full cost breakdowns"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Projekt Budget"
        
        # Enhanced headers with cost structure
        headers = [
            "Kategori", "Opgave", "Beskrivelse", "Kostpris_EP", "Materialer", 
            "Timer", "Takst", "Påslag_MAT", "Salgspris_MAT", "Påslag_UE", "Salgspris_UE", "Tilbud"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add project info
        ws.cell(row=2, column=1, value="PROJEKT INFO")
        ws.cell(row=3, column=1, value="Beskrivelse")
        ws.cell(row=3, column=2, value=project_data.description)
        ws.cell(row=4, column=1, value="Kvadratmeter")
        ws.cell(row=4, column=2, value=project_data.square_meters)
        
        # Add components with full cost breakdown
        current_row = 6
        total_cost = 0
        
        for category, components in project_data.selected_components.items():
            for component in components:
                # Extract cost data from component
                ws.cell(row=current_row, column=1, value=component.get('kategori', category))
                ws.cell(row=current_row, column=2, value=component.get('Opgave', ''))
                ws.cell(row=current_row, column=3, value=f"Komponent fra {component.get('source_file', 'videnbase')}")
                ws.cell(row=current_row, column=4, value=component.get('Kostpris_EP', 0))
                ws.cell(row=current_row, column=5, value=component.get('Materialer', 0))
                ws.cell(row=current_row, column=6, value=component.get('Timer', 0))
                ws.cell(row=current_row, column=7, value=component.get('Takst', 0))
                ws.cell(row=current_row, column=8, value=component.get('Påslag_MAT', 0))
                ws.cell(row=current_row, column=9, value=component.get('Salgspris_MAT', 0))
                ws.cell(row=current_row, column=10, value=component.get('Påslag_UE', 0))
                ws.cell(row=current_row, column=11, value=component.get('Salgspris_UE', 0))
                ws.cell(row=current_row, column=12, value=component.get('Tilbud', 0))
                
                total_cost += component.get('Tilbud', 0)
                current_row += 1
        
        # Add total row
        if current_row > 6:
            ws.cell(row=current_row, column=1, value="TOTAL")
            ws.cell(row=current_row, column=12, value=total_cost)
            
            # Style total row
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            total_font = Font(bold=True)
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = total_fill
                cell.font = total_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes for download
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, f"enhanced_projekt_budget_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
    except Exception as e:
        st.error(f"Fejl ved oprettelse af Excel-fil: {e}")
        return None, None

def run_enhanced_agent_ui():
    """Enhanced Streamlit UI for the AI agent with knowledge base integration"""
    st.markdown('<h1 class="main-header">🤖 Enhanced AI Agent med Videnbase Integration</h1>', unsafe_allow_html=True)
    
    # Check OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        st.error("❌ OPENAI_API_KEY miljøvariabel er ikke sat. Sæt venligst din OpenAI API nøgle.")
        st.info("Du kan sætte den i PowerShell med: $env:OPENAI_API_KEY='din_nøgle_her'")
        return
    
    # Initialize session state
    if 'project_data' not in st.session_state:
        st.session_state.project_data = ProjectData()
    
    if 'search_system' not in st.session_state:
        try:
            # Initialize search system with correct path
            kb_path = os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'unified_knowledge_base.json')
            from component_embeddings import ComponentEmbeddings
            
            # Create embeddings manager with correct path
            embeddings_manager = ComponentEmbeddings(knowledge_base_path=kb_path)
            
            # Create semantic search with the embeddings manager
            search_system = SemanticSearch(embeddings_manager=embeddings_manager)
            
            st.session_state.search_system = search_system
            st.success(f"✅ Søgesystem initialiseret med videnbase: {os.path.basename(kb_path)}")
        except Exception as e:
            st.error(f"❌ Kunne ikke initialisere søgesystem: {e}")
            st.session_state.search_system = None
    
    # Step 1: Project Description
    st.header("📝 Trin 1: Projektbeskrivelse")
    
    project_description = st.text_area(
        "Beskriv dit projekt",
        value=st.session_state.project_data.description,
        placeholder="F.eks. 'Total renovation af badeværelse med nye fliser og VVS'",
        height=100
    )
    
    if project_description:
        st.session_state.project_data.description = project_description
        
        # Get square meters
        col1, col2 = st.columns(2)
        with col1:
            square_meters = st.number_input(
                "Kvadratmeter",
                min_value=0.0,
                max_value=1000.0,
                value=st.session_state.project_data.square_meters,
                step=0.5,
                help="Indtast projektets størrelse i kvadratmeter"
            )
            st.session_state.project_data.square_meters = square_meters
        
        with col2:
            st.metric("Projekt", project_description[:30] + "..." if len(project_description) > 30 else project_description)
            st.metric("Størrelse", f"{square_meters} m²")
    
    # Step 2: Category Selection
    if st.session_state.project_data.description:
        st.header("🏗️ Trin 2: Vælg Relevante Kategorier")
        
        categories = load_categories()
        
        # Display categories with descriptions
        st.info("Vælg de kategorier der er relevante for dit projekt:")
        
        selected_categories = []
        cols = st.columns(3)
        
        for i, category in enumerate(categories):
            col_idx = i % 3
            with cols[col_idx]:
                if st.checkbox(f"**{category.title()}**", key=f"cat_{category}"):
                    selected_categories.append(category)
        
        st.session_state.project_data.selected_categories = selected_categories
        
        if selected_categories:
            st.success(f"✅ Valgt {len(selected_categories)} kategorier: {', '.join(selected_categories)}")
    
    # Step 3: Task Gathering with Knowledge Base Search
    if st.session_state.project_data.selected_categories:
        st.header("🔨 Trin 3: Opgaver med Omkostninger fra Videnbase")
        
        # Initialize components if not exists
        if 'selected_components' not in st.session_state.project_data.__dict__:
            st.session_state.project_data.selected_components = {}
        
        for category in st.session_state.project_data.selected_categories:
            st.markdown(f'<div class="category-section">', unsafe_allow_html=True)
            st.subheader(f"📋 {category.title()}")
            
            # Initialize components list for this category
            if category not in st.session_state.project_data.selected_components:
                st.session_state.project_data.selected_components[category] = []
            
            # Task input
            task_input = st.text_input(
                f"Hvilke underopgaver er nødvendige i {category}?",
                key=f"task_{category}",
                placeholder=f"F.eks. 'nedrivning af fliser' for {category}"
            )
            
            if task_input:
                # Search knowledge base
                st.info(f"🔍 Søger i videnbase efter: '{task_input}'")
                
                # Check if search system is available
                if not st.session_state.search_system:
                    st.error("⚠️ Søgesystemet er ikke tilgængeligt. Prøv at genindlæse siden.")
                    continue
                
                try:
                    search_results = st.session_state.search_system.search(task_input, top_k=5, min_similarity=0.2)
                    
                    if search_results:
                        st.success(f"Fandt {len(search_results)} relevante komponenter!")
                        
                        # Display search results
                        for i, result in enumerate(search_results):
                            with st.expander(f"🔍 Resultat {i+1}: {result.get('Opgave', 'N/A')[:50]}..."):
                                st.markdown(f"""
                                **Opgave:** {result.get('Opgave', 'N/A')}
                                **Kategori:** {result.get('kategori', 'N/A')}
                                **Lighedsscore:** {result.get('similarity_score', 0):.3f}
                                """)
                                
                                # Cost breakdown
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Total Pris", f"{result.get('Tilbud', 0):,.0f} DKK")
                                    st.metric("Kostpris EP", f"{result.get('Kostpris_EP', 0):,.0f} DKK")
                                with col2:
                                    st.metric("Materialer", f"{result.get('Materialer', 0):,.0f} DKK")
                                    st.metric("Timer", f"{result.get('Timer', 0):.1f}")
                                with col3:
                                    st.metric("Takst", f"{result.get('Takst', 0):,.0f} DKK")
                                    st.metric("Påslag MAT", f"{result.get('Påslag_MAT', 0):,.0f} DKK")
                                
                                # Add to project button
                                if st.button(f"Tilføj til projekt", key=f"add_{category}_{i}"):
                                    st.session_state.project_data.selected_components[category].append(result)
                                    st.success(f"✅ Tilføjet: {result.get('Opgave', 'N/A')}")
                                    st.rerun()
                        
                    else:
                        st.warning("Ingen komponenter fundet. Prøv at beskrive opgaven mere specifikt.")
                
                except Exception as e:
                    st.error(f"Fejl ved søgning: {e}")
            
            # Show selected components for this category
            if st.session_state.project_data.selected_components.get(category):
                st.subheader(f"✅ Valgte komponenter for {category}")
                for i, component in enumerate(st.session_state.project_data.selected_components[category]):
                    col1, col2, col3 = st.columns([3, 2, 1])
                    with col1:
                        st.write(f"**{component.get('Opgave', 'N/A')}**")
                    with col2:
                        st.metric("Pris", f"{component.get('Tilbud', 0):,.0f} DKK")
                    with col3:
                        if st.button("🗑️", key=f"remove_{category}_{i}"):
                            st.session_state.project_data.selected_components[category].pop(i)
                            st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Step 4: Project Summary and Excel Export
    if st.session_state.project_data.selected_categories:
        st.header("💾 Trin 4: Projekt Sammendrag og Excel Export")
        
        # Calculate totals
        total_components = sum(len(components) for components in st.session_state.project_data.selected_components.values())
        total_cost = sum(
            sum(comp.get('Tilbud', 0) for comp in components)
            for components in st.session_state.project_data.selected_components.values()
        )
        
        # Display summary
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Kategorier", len(st.session_state.project_data.selected_categories))
            st.metric("Total Komponenter", total_components)
        with col2:
            st.metric("Total Omkostning", f"{total_cost:,.0f} DKK")
            st.metric("Gns. per Komponent", f"{total_cost/total_components:,.0f} DKK" if total_components > 0 else "0 DKK")
        with col3:
            st.metric("Projekt Størrelse", f"{st.session_state.project_data.square_meters} m²")
            st.metric("Pris per m²", f"{total_cost/st.session_state.project_data.square_meters:,.0f} DKK" if st.session_state.project_data.square_meters > 0 else "0 DKK")
        
        # Excel export
        if total_components > 0:
            st.subheader("📊 Excel Export med Fuld Omkostningsstruktur")
            
            excel_data, filename = create_enhanced_excel_download(st.session_state.project_data)
            
            if excel_data:
                st.download_button(
                    label="📥 Download Enhanced Excel Budget",
                    data=excel_data.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download Excel-fil med fuld omkostningsstruktur inkl. Kostpris_EP, Materialer, Timer, Takst, Påslag og Salgspris felter"
                )
                
                st.success(f"✅ Excel-fil klar til download: {filename}")
                st.info("📋 Filen indeholder alle omkostningsdetaljer fra videnbasen med professionel formatering.")
            else:
                st.error("❌ Kunne ikke oprette Excel-fil")
        else:
            st.warning("⚠️ Ingen komponenter valgt endnu. Tilføj komponenter fra søgeresultaterne for at kunne eksportere til Excel.")

def main():
    """Main function"""
    st.sidebar.title("🏗️ Enhanced Construction Agent")
    
    # Sidebar navigation
    page = st.sidebar.radio(
        "Vælg agent:",
        ["Enhanced AI Agent", "Simple Agent", "AI Agent (Basic)"]
    )
    
    # Sidebar info
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🔍 Videnbase Integration")
    st.sidebar.markdown("""
    - **Semantic Search** gennem 633 komponenter
    - **Omkostningsstruktur** bevares i Excel
    - **AI-drevet** komponentvalg
    - **Professionel** budgetformatering
    """)
    
    st.sidebar.markdown("### 📊 Omkostningsfelter")
    st.sidebar.markdown("""
    - Kostpris_EP
    - Materialer
    - Timer & Takst
    - Påslag_MAT/UE
    - Salgspris_MAT/UE
    - Total Tilbud
    """)
    
    # Check OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        st.sidebar.error("❌ OPENAI_API_KEY mangler")
        st.sidebar.info("Sæt miljøvariabel for at bruge AI-funktioner")
    
    # Page routing
    if page == "Enhanced AI Agent":
        run_enhanced_agent_ui()
    elif page == "Simple Agent":
        st.info("🚧 Simple Agent UI kommer snart...")
        st.write("Brug 'Enhanced AI Agent' for fuld funktionalitet med videnbase integration.")
    elif page == "AI Agent (Basic)":
        st.info("🚧 Basic AI Agent UI kommer snart...")
        st.write("Brug 'Enhanced AI Agent' for fuld funktionalitet med videnbase integration.")

if __name__ == "__main__":
    main()
