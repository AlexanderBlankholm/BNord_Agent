import os
import json
from typing import Dict, List, Any
from dataclasses import dataclass
from datetime import datetime
import sys

# Add the knowledge_base directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'embeddings_and_search'))

# LangChain imports
from langchain.agents import initialize_agent, AgentType, Tool
from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.memory import ConversationBufferMemory
from langchain.chains import LLMChain
from langchain.schema import HumanMessage, AIMessage

# Import the existing search system
from semantic_search import SemanticSearch

@dataclass
class ProjectData:
    """Enhanced data structure to hold project information with cost details"""
    description: str = ""
    detailed_description: str = ""
    selected_categories: List[str] = None
    category_tasks: Dict[str, List[str]] = None
    selected_components: Dict[str, List[Dict[str, Any]]] = None  # Store selected components with costs
    
    def __post_init__(self):
        if self.selected_categories is None:
            self.selected_categories = []
        if self.category_tasks is None:
            self.category_tasks = {}
        if self.selected_components is None:
            self.selected_components = {}

class EnhancedAIAgent:
    """Enhanced AI-powered agent using LangChain and OpenAI with knowledge base integration"""
    
    def __init__(self):
        self.project_data = ProjectData()
        self.categories = self.load_categories()
        self.memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
        
        # Initialize OpenAI LLM (using GPT-3.5-turbo for cost-efficient testing)
        self.llm = ChatOpenAI(
            temperature=0.1,  # Low temperature for consistent responses
            model="gpt-3.5-turbo"  # Cost-efficient model for testing phase
        )
        
        # Initialize knowledge base search system with correct path
        # The search system needs to know where the knowledge base is located
        self.search_system = self._initialize_search_system()
        
        # Define enhanced tools
        self.tools = [
            Tool(
                name="SearchKnowledgeBase",
                func=self.search_knowledge_base,
                description="Search the construction knowledge base for components matching a task description. Use this to find relevant components with cost breakdowns."
            ),
            Tool(
                name="GenerateComponent",
                func=self.generate_component_with_rag,
                description="Generate a new construction component using RAG pipeline based on user query and matched components from knowledge base."
            ),
            Tool(
                name="GetCategories",
                func=self.get_categories,
                description="Returns the list of available construction categories"
            ),
            Tool(
                name="SaveToExcel",
                func=self.save_to_excel,
                description="Saves the final project data to an Excel file with full cost breakdowns"
            )
        ]
        
        # Initialize the agent
        self.agent = initialize_agent(
            tools=self.tools,
            llm=self.llm,
            agent=AgentType.CONVERSATIONAL_REACT_DESCRIPTION,
            verbose=True,
            memory=self.memory,
            handle_parsing_errors=True
        )
        
        # Create specialized prompts for different stages
        self.create_prompts()
    
    def load_categories(self) -> List[str]:
        """Load categories from JSON file"""
        try:
            with open("categories.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("categories", [])
        except FileNotFoundError:
            print("Warning: categories.json not found, using default categories")
            return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft"]
    
    def create_prompts(self):
        """Create specialized prompts for different conversation stages"""
        
        # Initial project description prompt
        self.initial_prompt = PromptTemplate(
            input_variables=["chat_history"],
            template="""Du er en ekspert i byggeprojekter og badeværelsesrenovering. Din opgave er at hjælpe brugeren med at planlægge deres projekt.

{chat_history}

Du skal starte med at spørge om projektets beskrivelse. Stil spørgsmål på dansk og vær venlig og professionel.

Hvis brugeren allerede har beskrevet deres projekt, så spørg om yderligere detaljer som kvadratmeter."""
        )
        
        # Follow-up questions prompt
        self.followup_prompt = PromptTemplate(
            input_variables=["chat_history", "project_description"],
            template="""Baseret på projektbeskrivelsen: "{project_description}"

{chat_history}

Stil relevante opfølgende spørgsmål på dansk for at få mere information om projektet. 
Spørg specifikt om kvadratmeter og andre vigtige detaljer."""
        )
        
        # Category selection prompt
        self.category_prompt = PromptTemplate(
            input_variables=["chat_history", "categories"],
            template="""Her er de tilgængelige byggekategorier:

{categories}

{chat_history}

Præsenter disse kategorier for brugeren på dansk og hjælp dem med at vælge hvilke der er relevante for deres projekt.
Forklar kort hvad hver kategori indebærer."""
        )
        
        # Task gathering prompt with knowledge base integration
        self.task_prompt = PromptTemplate(
            input_variables=["chat_history", "category"],
            template="""Nu arbejder vi med kategorien: {category}

{chat_history}

Spørg brugeren om hvilke specifikke underopgaver der er nødvendige i denne kategori.
Giv eksempler på typiske opgaver for {category}.
Når brugeren beskriver en opgave, brug SearchKnowledgeBase værktøjet til at finde relevante komponenter med omkostninger.
Husk at spørge om der er flere opgaver før du går videre til næste kategori."""
        )
    
    def search_knowledge_base(self, query: str) -> str:
        """Tool function to search the knowledge base for components"""
        try:
            # Check if search system is available
            if not self.search_system:
                return f"⚠️ Søgesystemet er ikke tilgængeligt. Prøv at genstarte agenten."
            
            # Search for components matching the query
            results = self.search_system.search(query, top_k=5, min_similarity=0.2)
            
            if not results:
                return f"Ingen komponenter fundet for '{query}'. Prøv at beskrive opgaven mere specifikt."
            
            # Format results for the agent
            formatted_results = []
            for result in results:
                opgave = result.get('Opgave', 'N/A')
                kategori = result.get('kategori', 'N/A')
                tilbud = result.get('Tilbud', 0)
                kostpris_ep = result.get('Kostpris_EP', 0)
                materialer = result.get('Materialer', 0)
                timer = result.get('Timer', 0)
                takst = result.get('Takst', 0)
                similarity = result.get('similarity_score', 0)
                
                formatted_results.append(
                    f"• {opgave} ({kategori})\n"
                    f"  - Total pris: {tilbud:,.0f} DKK\n"
                    f"  - Kostpris EP: {kostpris_ep:,.0f} DKK\n"
                    f"  - Materialer: {materialer:,.0f} DKK\n"
                    f"  - Timer: {timer:.1f} | Takst: {takst:,.0f} DKK\n"
                    f"  - Lighedsscore: {similarity:.3f}"
                )
            
            return f"Fandt {len(results)} relevante komponenter for '{query}':\n\n" + "\n\n".join(formatted_results)
            
        except Exception as e:
            return f"Fejl ved søgning i videnbasen: {str(e)}"
    
    def get_categories(self) -> str:
        """Tool function to get available categories"""
        return f"Tilgængelige kategorier: {', '.join(self.categories)}"
    
    def save_to_excel(self, project_summary: str) -> str:
        """Enhanced tool function to save project data to Excel with full cost breakdowns"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Projekt Budget"
            
            # Enhanced headers with cost structure
            headers = [
                "Kategori", "Opgave", "Beskrivelse", "Admin", "Kostpris_EP", "Materialer", 
                "Timer", "Takst", "Påslag_MAT", "Salgspris_MAT", "Påslag_UE", "Salgspris_UE", "Tilbud"
            ]
            
            # Style for headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add project info
            ws.cell(row=2, column=1, value="PROJEKT INFO")
            ws.cell(row=3, column=1, value="Titel")
            ws.cell(row=3, column=2, value=self.project_data.description)
            ws.cell(row=4, column=1, value="Beskrivelse")
            ws.cell(row=4, column=2, value=self.project_data.detailed_description)
            
            # Add components with full cost breakdown
            current_row = 6
            total_cost = 0
            
            for category, components in self.project_data.selected_components.items():
                for component in components:
                    # Extract cost data from component
                    ws.cell(row=current_row, column=1, value=component.get('kategori', category))
                    ws.cell(row=current_row, column=2, value=component.get('Opgave', ''))
                    ws.cell(row=current_row, column=3, value=f"Komponent fra {component.get('source_file', 'videnbase')}")
                    ws.cell(row=current_row, column=4, value=component.get('Admin', 0)) # Add Admin field
                    ws.cell(row=current_row, column=5, value=component.get('Kostpris_EP', 0))
                    ws.cell(row=current_row, column=6, value=component.get('Materialer', 0))
                    ws.cell(row=current_row, column=7, value=component.get('Timer', 0))
                    ws.cell(row=current_row, column=8, value=component.get('Takst', 0))
                    ws.cell(row=current_row, column=9, value=component.get('Påslag_MAT', 0))
                    ws.cell(row=current_row, column=10, value=component.get('Salgspris_MAT', 0))
                    ws.cell(row=current_row, column=11, value=component.get('Påslag_UE', 0))
                    ws.cell(row=current_row, column=12, value=component.get('Salgspris_UE', 0))
                    ws.cell(row=current_row, column=13, value=component.get('Tilbud', 0))
                    
                    total_cost += component.get('Tilbud', 0)
                    current_row += 1
            
            # Add total row
            if current_row > 6:
                ws.cell(row=current_row, column=1, value="TOTAL")
                ws.cell(row=current_row, column=13, value=total_cost) # Adjust column for total cost
                
                # Style total row
                total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                total_font = Font(bold=True)
                for col in range(1, 14): # Adjust range for total row
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = total_fill
                    cell.font = total_font
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"enhanced_projekt_budget_{timestamp}.xlsx"
            wb.save(filename)
            
            return f"Excel-fil gemt som: {filename} med total omkostning: {total_cost:,.0f} DKK"
            
        except Exception as e:
            return f"Fejl ved oprettelse af Excel-fil: {e}"
    
    def run_conversation(self):
        """Enhanced conversation loop using AI agent with knowledge base integration"""
        print("🤖 Hej! Jeg er din forbedrede AI-assistent til at planlægge dit byggeprojekt.")
        print("Jeg kan nu søge i videnbasen og give dig præcise omkostninger!")
        print("=" * 70)
        
        # Step 1: Get project description
        print("\n📝 **Trin 1: Projektbeskrivelse**")
        user_input = input("Beskriv dit projekt (f.eks. 'Total renovation af badeværelse'): ")
        self.project_data.description = user_input
        
        # Use AI to ask follow-up questions
        response = self.agent.run(f"Brugeren har beskrevet deres projekt som: {user_input}. Stil relevante opfølgende spørgsmål på dansk.")
        print(f"\n🤖 {response}")
        
        # Get detailed project description
        detailed_description = input("Beskriv dit projekt i detaljer (tryk Enter for at springe over): ")
        self.project_data.detailed_description = detailed_description
        
        # Step 2: Present categories with AI assistance
        print(f"\n🏗️ **Trin 2: Kategorier**")
        categories_text = "\n".join([f"{i+1}. {cat}" for i, cat in enumerate(self.categories)])
        
        response = self.agent.run(f"Præsenter disse kategorier for brugeren og forklar hvilke der kunne være relevante for et projekt om: {self.project_data.description}")
        print(f"\n🤖 {response}")
        
        # User selects categories
        print(f"\n{categories_text}")
        selection = input("Indtast numrene på de relevante kategorier (adskilt med komma, f.eks. 1,3,5): ")
        
        try:
            selected_indices = [int(x.strip()) - 1 for x in selection.split(",")]
            self.project_data.selected_categories = [self.categories[idx] for idx in selected_indices if 0 <= idx < len(self.categories)]
        except ValueError:
            self.project_data.selected_categories = []
        
        # Step 3: AI-guided task gathering with knowledge base search
        print(f"\n🔨 **Trin 3: Opgaver for hver kategori med omkostninger**")
        
        for category in self.project_data.selected_categories:
            print(f"\n📋 **Kategori: {category}**")
            
            # Initialize components list for this category
            if category not in self.project_data.selected_components:
                self.project_data.selected_components[category] = []
            
            # Use AI to ask about tasks and search knowledge base
            response = self.agent.run(f"Spørg brugeren om hvilke specifikke opgaver der er nødvendige i kategorien '{category}' for et projekt om {self.project_data.description}. Når de beskriver en opgave, brug SearchKnowledgeBase værktøjet til at finde relevante komponenter.")
            print(f"\n🤖 {response}")
            
            tasks = []
            while True:
                task = input(f"Hvilke underopgaver er nødvendige i {category}? ")
                if task.strip():
                    tasks.append(task.strip())
                    
                    # Search knowledge base for this task
                    print(f"\n🔍 Søger i videnbasen efter: {task}")
                    search_result = self.search_knowledge_base(task)
                    print(f"\n{search_result}")
                    
                    # Ask user to select components
                    print(f"\nVil du tilføje komponenter fra søgeresultaterne til dit projekt? (ja/nej): ")
                    add_components = input().lower().strip()
                    
                    if add_components in ['ja', 'j', 'yes', 'y']:
                        # Add components from search results
                        if self.search_system:
                            search_results = self.search_system.search(task, top_k=3, min_similarity=0.2)
                            if search_results:
                                print(f"\nVælg komponenter at tilføje (indtast numre adskilt med komma, f.eks. 1,2):")
                                for i, result in enumerate(search_results, 1):
                                    opgave = result.get('Opgave', 'N/A')
                                    tilbud = result.get('Tilbud', 0)
                                    print(f"{i}. {opgave} - {tilbud:,.0f} DKK")
                                
                                selection = input("Vælg komponenter: ").strip()
                                try:
                                    selected_indices = [int(x.strip()) - 1 for x in selection.split(",")]
                                    for idx in selected_indices:
                                        if 0 <= idx < len(search_results):
                                            component = search_results[idx]
                                            self.project_data.selected_components[category].append(component)
                                            print(f"✅ Tilføjet: {component.get('Opgave', 'N/A')}")
                                except ValueError:
                                    print("Ugyldig valg, ingen komponenter tilføjet")
                            else:
                                print("Ingen komponenter fundet at tilføje")
                        else:
                            print("⚠️ Søgesystemet er ikke tilgængeligt")
                
                more = input("Er der andet du vil tilføje til denne kategori? (ja/nej): ").lower().strip()
                if more not in ['ja', 'j', 'yes', 'y']:
                    break
            
            self.project_data.category_tasks[category] = tasks
            print(f"✅ {len(tasks)} opgaver tilføjet til {category}")
        
        # Step 4: AI generates project summary and saves to Excel
        print(f"\n💾 **Trin 4: Gem til Excel med fuld omkostningsstruktur**")
        
        # Create project summary
        total_components = sum(len(components) for components in self.project_data.selected_components.values())
        total_cost = sum(
            sum(comp.get('Tilbud', 0) for comp in components)
            for components in self.project_data.selected_components.values()
        )
        
        summary = f"""
Projekt: {self.project_data.description}
        Beskrivelse: {self.project_data.detailed_description[:100] + '...' if len(self.project_data.detailed_description) > 100 else self.project_data.detailed_description}
Kategorier: {', '.join(self.project_data.selected_categories)}
Total opgaver: {sum(len(tasks) for tasks in self.project_data.category_tasks.values())}
Valgte komponenter: {total_components}
Total omkostning: {total_cost:,.0f} DKK
        """.strip()
        
        response = self.agent.run(f"Opret et sammendrag af projektet og brug SaveToExcel værktøjet til at gemme dataene med fuld omkostningsstruktur. Projekt sammendrag: {summary}")
        print(f"\n🤖 {response}")
        
        print(f"\n✅ Projektet er nu færdigt! Excel-filen er oprettet med fuld omkostningsstruktur.")

    def generate_component_with_rag(self, user_query: str, use_high_quality_only: bool = True) -> Dict[str, Any]:
        """
        Generate a new construction component using RAG pipeline.
        Uses the user query and matched components from knowledge base to generate
        a high-quality component with proper pricing structure.
        
        Args:
            user_query: User's task description
            use_high_quality_only: If True, only use high-quality components for context
        """
        print(f"🚀 DEBUG: generate_component_with_rag called with query: '{user_query}'")
        try:
            print(f"🚀 DEBUG: Starting search for components...")
            # First, search for similar components to use as context
            if not self.search_system:
                return {"error": "Search system not available"}
            
            # Step 1: Try to find high-quality components first (from the three detailed source files)
            high_quality_results = self.search_system.search(user_query, top_k=5, min_similarity=0.2, min_quality_score=1.0)
            
            if use_high_quality_only and high_quality_results:
                # Use only high-quality components
                context_components = self._prepare_context_components(high_quality_results)
                context_quality = "high"
                context_source = f"high-quality ({len(high_quality_results)} components)"
                
            elif use_high_quality_only and not high_quality_results:
                # Fallback to full database when no high-quality matches found
                print("⚠️ No high-quality components found - using broader database for context")
                all_results = self.search_system.search(user_query, top_k=5, min_similarity=0.2, min_quality_score=0.0)
                if all_results:
                    context_components = self._prepare_context_components(all_results)
                    context_quality = "mixed"
                    context_source = f"full database ({len(all_results)} components, quality may vary)"
                else:
                    return {"error": "No similar components found to base generation on"}
                    
            else:
                # User chose to include all components
                all_results = self.search_system.search(user_query, top_k=5, min_similarity=0.2, min_quality_score=0.0)
                if all_results:
                    context_components = self._prepare_context_components(all_results)
                    context_quality = "mixed"
                    context_source = f"full database ({len(all_results)} components, quality may vary)"
                else:
                    return {"error": "No similar components found to base generation on"}
            
            # Debug: Print what the LLM will see
            print(f"🔍 DEBUG: Context components being sent to LLM:")
            for i, comp in enumerate(context_components):
                print(f"  Component {i+1}: Admin={comp.get('Admin', 'NOT_FOUND')}, Tilbud={comp.get('Tilbud', 'NOT_FOUND')}")
            
            # Create intelligent prompt for component generation with business rule compliance
            generation_prompt = f"""
            Generate a construction component based on: "{user_query}"
            
            Context components: {json.dumps(context_components, indent=2, ensure_ascii=False)}
            
            ## 🧠 INTELLIGENT COMPONENT GENERATION
            
            **APPROACH**: Be creative and intelligent! Use context components as inspiration and patterns to learn from, NOT as templates to copy from.
            
            **CONTEXT ANALYSIS**:
            - Study the most similar component as your starting point
            - Analyze ALL retrieved components to understand realistic value ranges
            - Apply business knowledge and construction expertise
            - Make intelligent estimates based on the task description
            
            ## 🎯 BUSINESS RULES (ALWAYS ENFORCE)
            
            **MATHEMATICAL RELATIONSHIPS (NEVER BREAK):**
            - Kostpris_EP = Timer × Takst (ALWAYS calculate this)
            - Salgspris_MAT = Materialer × (1 + Påslag_MAT/100) (ALWAYS calculate this)
            - Salgspris_UE = UE × (1 + Påslag_UE/100) (ALWAYS calculate this)
            - Tilbud = Admin + Kostpris_EP + Salgspris_MAT + Salgspris_UE (ALWAYS calculate this)
            
            **TRADE-SPECIFIC RULES:**
            - **Fag = "Bnord"**: UE = 0, Påslag_UE = 0, Salgspris_UE = 0 (no subcontractor costs)
            - **Fag ≠ "Bnord"**: Can have UE costs for subcontractor work
            
            **CATEGORY-SPECIFIC RULES:**
            - **Kategori = "Projekt"**: 
              - **EXCEPTION**: If query contains 'afhentning', 'affald', 'garbage' → treat as flat fee service
                - Admin = 0, Timer = 0, Takst = 0, Kostpris_EP = 0, Materialer = 0
                - Tilbud = [estimate flat fee based on context]
              - **STANDARD**: Only Admin costs, all other fields = 0
            
            ## 🏗️ COMPONENT TYPE ANALYSIS
            
            **Analyze the task and context to determine:**
            - **Labor requirements**: How many hours (Timer) and what rate (Takst)?
            - **Material requirements**: What materials are needed and their cost?
            - **Markup strategy**: What are realistic markup percentages for this type of work?
            - **Admin overhead**: Is this a project management task requiring Admin costs?
            - **Service type**: Is this a flat fee service (like garbage collection)?
            
            **Use context components to understand:**
            - Realistic time estimates for similar tasks
            - Typical material costs and markup patterns
            - Hourly rates for different types of work
            - Value ranges that make business sense
            - Flat fee ranges for service-type work
            
            ## 📊 FIELD ASSIGNMENT STRATEGY
            
            **INTELLIGENT ESTIMATION PROCESS:**
            1. **Study the most similar context component** as your starting inspiration
            2. **Analyze all context components** to understand realistic value ranges
            3. **Apply business knowledge** about construction tasks, materials, and labor
            4. **Make educated guesses** for each field based on the task description
            5. **Ensure mathematical consistency** by calculating derived fields
            6. **Validate against business rules** (Bnord vs. subcontractor work)
            
            **NEVER**: Strictly copy values from context components
            **ALWAYS**: Use context as inspiration to make intelligent, realistic estimates
            
            ## 💰 DEFAULT HOURLY RATES (Takst)
            
            **Use these as preferred values unless context strongly suggests otherwise:**
            - **Nedrivning category**: 500 DKK/hour (demolition work typically has lower rates)
            - **All other categories**: 585 DKK/hour (standard construction labor rate)
            
            **Context components may show different rates** - use them as guidance but prefer the defaults unless there's a compelling reason (e.g., specialized skills, high-complexity work, or consistent patterns in similar components).
            
            ## 🗑️ SPECIAL CASE: GARBAGE COLLECTION SERVICES
            
            **If query contains 'afhentning', 'affald', 'garbage', 'big bag':**
            - **Treat as flat fee service** (not project management)
            - **Admin = 0** (no project overhead)
            - **Timer = 0, Takst = 0, Kostpris_EP = 0** (no labor costs)
            - **Materialer = 0** (no materials)
            - **Tilbud = [estimate based on context ranges]** (direct service cost)
            - **Example**: If context shows 5,850-7,020 DKK, estimate around 6,500 DKK
            
            ## 💰 MARKUP GUIDANCE
            
            **Material Markup (Påslag_MAT):**
            - **Preferred range**: 17-20% (industry standard for construction materials)
            - **If context shows unrealistic values** (0%, very low %): Prefer 17-20%
            - **Apply business knowledge**: Materials typically need markup for handling, storage, profit
            - **Use context as reference** but prioritize realistic business practices
            
            **Subcontractor Markup (Påslag_UE):**
            - **Typical range**: 15-25% (depends on trade and complexity)
            - **Only for non-Bnord work** (subcontractor services)
            
            ## 🚀 GENERATION INSTRUCTIONS
            
            Generate a JSON component with:
            - Opgave: "{user_query}"
            - kategori: Determine based on task type and context
            - Fag: "Bnord" for in-house work, other trades for subcontractor work
            - Timer: Estimate based on context and business knowledge
            - Takst: Estimate based on context and business knowledge  
            - Materialer: Estimate based on context and business knowledge
            - Påslag_MAT: Estimate based on context and business knowledge (prefer 17-20%)
            - Admin: Estimate if project management overhead is needed
            - UE: 0 for Bnord work, estimate for subcontractor work
            - Påslag_UE: 0 for Bnord work, estimate for subcontractor work
            
            **CRITICAL**: Calculate Kostpris_EP, Salgspris_MAT, Salgspris_UE, and Tilbud using the mathematical formulas above!
            
            **REMEMBER**: Be creative, intelligent, and business-savvy while maintaining consistency!
            """
            
            # Use the LLM to generate the component
            print(f"🚀 DEBUG: About to call LLM...")
            print(f"🚀 DEBUG: Prompt length: {len(generation_prompt)} characters")
            
            try:
                response = self.llm.invoke(generation_prompt)
                print(f"🚀 DEBUG: LLM call completed successfully")
                print(f"🔍 Debug: LLM response type: {type(response)}")
                print(f"🔍 Debug: LLM response content: {response.content}")
                print(f"🔍 Debug: Response content length: {len(response.content) if response.content else 0}")
                
                # Check if response is empty
                if not response.content or len(response.content.strip()) == 0:
                    print(f"🚨 ERROR: LLM returned empty response!")
                    print(f"🚨 ERROR: This usually means the prompt was too long or there's an API issue")
                    raise Exception("LLM returned empty response - prompt may be too long or API issue")
                
                print(f"🚀 DEBUG: About to parse JSON...")
                
                # Try to parse the response as JSON
                try:
                    # Extract JSON content from markdown response
                    content = response.content.strip()
                    print(f"🔍 DEBUG: Raw content: '{content}'")
                    
                    if content.startswith('```json'):
                        # Remove markdown code block markers
                        json_start = content.find('{')
                        json_end = content.rfind('}') + 1
                        if json_start != -1 and json_end != 0:
                            json_content = content[json_start:json_end]
                            print(f"🚀 DEBUG: Extracted JSON content: {json_content}")
                            generated_component = json.loads(json_content)
                        else:
                            raise json.JSONDecodeError("Could not find JSON content in markdown", content, 0)
                    else:
                        # Try to parse directly
                        generated_component = json.loads(content)
                    
                    print(f"🚀 DEBUG: JSON parsing completed successfully")
                except json.JSONDecodeError as e:
                    print(f"🚨 DEBUG: JSON parsing failed: {e}")
                    print(f"🚨 DEBUG: Failed content: '{response.content}'")
                    print(f"🚨 DEBUG: Content type: {type(response.content)}")
                    print(f"🚨 DEBUG: Content length: {len(response.content) if response.content else 0}")
                    raise
                except Exception as e:
                    print(f"🚨 DEBUG: Unexpected error during JSON parsing: {e}")
                    raise
                print(f"🔍 Debug: Successfully parsed JSON, Tilbud = {generated_component.get('Tilbud', 'NOT_FOUND')}")
                print(f"🔍 Debug: Initial component state:")
                print(f"   UE: {generated_component.get('UE', 'NOT_FOUND')}")
                print(f"   Påslag_UE: {generated_component.get('Påslag_UE', 'NOT_FOUND')}")
                print(f"   Tilbud: {generated_component.get('Tilbud', 'NOT_FOUND')}")
                
                # Validate and enforce business rules and mathematical relationships
                print(f"🔍 Debug: Validating and enforcing business rules...")
                
                # ENFORCE: Mathematical relationships (ALWAYS)
                if 'Timer' in generated_component and 'Takst' in generated_component:
                    calculated_kostpris = generated_component['Timer'] * generated_component['Takst']
                    generated_component['Kostpris_EP'] = calculated_kostpris
                    print(f"✅ Calculated Kostpris_EP: {generated_component['Timer']} × {generated_component['Takst']} = {calculated_kostpris}")
                
                # ENFORCE: Material markup logic - only apply markup if there are materials
                if generated_component.get('Materialer', 0) == 0:
                    print(f"✅ No materials used - setting material markup and sales price to 0")
                    generated_component['Påslag_MAT'] = 0
                    generated_component['Salgspris_MAT'] = 0
                else:
                    # ENFORCE: Realistic markup guidance (only for components with materials)
                    if generated_component.get('Påslag_MAT', 0) < 15:
                        print(f"⚠️ Material markup too low ({generated_component['Påslag_MAT']}%). Adjusting to realistic 17%...")
                        generated_component['Påslag_MAT'] = 17.0
                    
                    # Ensure Salgspris_MAT is calculated correctly with current markup
                    current_markup = generated_component.get('Påslag_MAT', 0)
                    material_cost = generated_component.get('Materialer', 0)
                    generated_component['Salgspris_MAT'] = material_cost * (1 + current_markup/100)
                    print(f"✅ Calculated Salgspris_MAT: {material_cost} × (1 + {current_markup}%) = {generated_component['Salgspris_MAT']}")
                
                if 'UE' in generated_component and 'Påslag_UE' in generated_component:
                    if generated_component['UE'] > 0:
                        påslag = generated_component['Påslag_UE'] / 100
                        generated_component['Salgspris_UE'] = generated_component['UE'] * (1 + påslag)
                        print(f"✅ Calculated Salgspris_UE: {generated_component['UE']} × (1 + {generated_component['Påslag_UE']}%) = {generated_component['Salgspris_UE']}")
                    else:
                        generated_component['Salgspris_UE'] = 0
                        print(f"✅ Salgspris_UE set to 0 (no UE costs)")
                
                # ENFORCE: Business rules for Bnord work
                if generated_component.get('Fag', '') == 'Bnord':
                    if generated_component.get('UE', 0) > 0:
                        print(f"⚠️ LLM incorrectly set UE > 0 for Bnord work. Fixing: UE {generated_component['UE']} → 0")
                        generated_component['UE'] = 0
                        generated_component['Påslag_UE'] = 0
                        generated_component['Salgspris_UE'] = 0
                    print(f"✅ Enforced Bnord rules: UE = 0, Påslag_UE = 0, Salgspris_UE = 0")
                
                # ENFORCE: Default hourly rates (Takst) based on category
                if generated_component.get('Timer', 0) > 0:  # Only apply if there are labor hours
                    current_takst = generated_component.get('Takst', 0)
                    kategori = generated_component.get('kategori', '')
                    
                    if kategori == 'Nedrivning':
                        default_takst = 500
                        if current_takst != default_takst:
                            print(f"⚠️ Adjusting Takst for Nedrivning category: {current_takst} → {default_takst} DKK/hour")
                            generated_component['Takst'] = default_takst
                        else:
                            print(f"✅ Takst correctly set for Nedrivning: {default_takst} DKK/hour")
                    else:
                        default_takst = 585
                        if current_takst != default_takst:
                            print(f"⚠️ Adjusting Takst to default rate: {current_takst} → {default_takst} DKK/hour")
                            generated_component['Takst'] = default_takst
                        else:
                            print(f"✅ Takst correctly set to default: {default_takst} DKK/hour")
                    
                    # Recalculate Kostpris_EP with potentially adjusted Takst
                    timer = generated_component.get('Timer', 0)
                    takst = generated_component.get('Takst', 0)
                    generated_component['Kostpris_EP'] = timer * takst
                    print(f"✅ Recalculated Kostpris_EP: {timer} × {takst} = {generated_component['Kostpris_EP']}")
                
                # ENFORCE: Project category rules with garbage collection exception
                if generated_component.get('kategori', '') in ['Projekt', 'Service']:
                    query_lower = user_query.lower()
                    is_garbage_collection = any(keyword in query_lower for keyword in ['afhentning', 'affald', 'garbage', 'big bag'])
                    
                    if is_garbage_collection:
                        print(f"✅ Detected garbage collection service - applying flat fee rules...")
                        generated_component['Admin'] = 0
                        generated_component['Timer'] = 0
                        generated_component['Takst'] = 0
                        generated_component['Kostpris_EP'] = 0
                        generated_component['Materialer'] = 0
                        generated_component['Påslag_MAT'] = 0
                        generated_component['Salgspris_MAT'] = 0
                        generated_component['UE'] = 0
                        generated_component['Påslag_UE'] = 0
                        generated_component['Salgspris_UE'] = 0
                        
                        # For garbage collection, preserve the estimated Tilbud from LLM
                        # but ensure it's not 0 (should be a realistic flat fee)
                        if generated_component.get('Tilbud', 0) == 0:
                            # If LLM didn't set a realistic Tilbud, estimate from context
                            context_tilbud_values = [c.get('Tilbud', 0) for c in context_components if c.get('Tilbud', 0) > 0]
                            if context_tilbud_values:
                                estimated_tilbud = sum(context_tilbud_values) / len(context_tilbud_values)
                                generated_component['Tilbud'] = round(estimated_tilbud, 0)
                                print(f"✅ Estimated Tilbud from context: {estimated_tilbud:.0f} DKK")
                        
                        print(f"✅ Garbage collection: Flat fee service, Tilbud = {generated_component.get('Tilbud', 0):,.0f} DKK")
                        
                        # Skip final Tilbud calculation for garbage collection (preserve estimated value)
                        skip_tilbud_calculation = True
                    else:
                        print(f"✅ Enforcing standard Projekt category rules...")
                        generated_component['Timer'] = 0
                        generated_component['Takst'] = 0
                        generated_component['Kostpris_EP'] = 0
                        generated_component['Materialer'] = 0
                        generated_component['Påslag_MAT'] = 0
                        generated_component['Salgspris_MAT'] = 0
                        generated_component['UE'] = 0
                        generated_component['Påslag_UE'] = 0
                        generated_component['Salgspris_UE'] = 0
                        print(f"✅ Projekt category: Only Admin costs allowed, all other fields = 0")
                        skip_tilbud_calculation = False
                else:
                    skip_tilbud_calculation = False
                
                # ENFORCE: Calculate final Tilbud (ALWAYS)
                if not skip_tilbud_calculation:
                    admin = generated_component.get('Admin', 0)
                    kostpris_ep = generated_component.get('Kostpris_EP', 0)
                    salgspris_mat = generated_component.get('Salgspris_MAT', 0)
                    salgspris_ue = generated_component.get('Salgspris_UE', 0)
                    
                    calculated_tilbud = admin + kostpris_ep + salgspris_mat + salgspris_ue
                    generated_component['Tilbud'] = calculated_tilbud
                    print(f"✅ Calculated Tilbud: {admin} + {kostpris_ep} + {salgspris_mat} + {salgspris_ue} = {calculated_tilbud}")
                
                # Ensure Admin field is present
                if 'Admin' not in generated_component:
                    generated_component['Admin'] = 0
                
                # Add context quality information
                generated_component['context_quality'] = context_quality
                generated_component['context_source'] = context_source
                
                # Store the actual context components used for UI display
                generated_component['context_components_used'] = [
                    {
                        'Opgave': c['Opgave'],
                        'kategori': c['kategori'],
                        'Fag': c['Fag'],
                        'Admin': c.get('Admin', 0),
                        'Kostpris_EP': c.get('Kostpris_EP', 0),
                        'Materialer': c.get('Materialer', 0),
                        'Timer': c.get('Timer', 0),
                        'Takst': c.get('Takst', 0),
                        'Tilbud': c['Tilbud'],
                        'similarity_score': c.get('similarity_score', 0.0),
                        'quality_score': c['quality_score'],
                        'source_file': c.get('source_file', 'N/A')
                    }
                    for c in context_components
                ]
                
                # Trust the LLM output - no more forced fallbacks
                print(f"🔍 Debug: Final check before return:")
                print(f"   Tilbud = {generated_component.get('Tilbud', 'NOT_FOUND')}")
                print(f"   UE = {generated_component.get('UE', 'NOT_FOUND')}")
                print(f"   Påslag_UE = {generated_component.get('Påslag_UE', 'NOT_FOUND')}")
                print(f"   Full component: {json.dumps(generated_component, indent=2, ensure_ascii=False)}")
                
                print(f"🔍 Debug: Returning component with Tilbud = {generated_component.get('Tilbud', 'NOT_FOUND')}")
                return generated_component
                
            except json.JSONDecodeError as e:
                # If LLM didn't return valid JSON, that's a problem - don't try to fix it
                print(f"❌ LLM failed to generate valid JSON: {e}")
                return {"error": f"LLM failed to generate valid component structure: {str(e)}"}
                
        except Exception as e:
            return {"error": f"Failed to generate component: {str(e)}"}

    def _prepare_context_components(self, search_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare context components for RAG generation"""
        context_components = []
        for result in search_results:
            # Keep original field names for better LLM understanding
            context_components.append({
                'Opgave': result.get('Opgave', ''),
                'kategori': result.get('kategori', ''),
                'Fag': result.get('Fag', ''),
                'Admin': result.get('Admin', 0),
                'Kostpris_EP': result.get('Kostpris_EP', 0),
                'Materialer': result.get('Materialer', 0),
                'Timer': result.get('Timer', 0),
                'Takst': result.get('Takst', 0),
                'Påslag_MAT': result.get('Påslag_MAT', 0),
                'Påslag_UE': result.get('Påslag_UE', 0),
                'UE': result.get('UE', 0),
                'Tilbud': result.get('Tilbud', 0),
                'quality_score': result.get('quality_score', 0.5),
                'source_file': result.get('source_file', 'N/A'),
                'similarity_score': result.get('similarity_score', 0.0)
            })
        return context_components

    def _get_component_quality_score(self, component: Dict[str, Any]) -> float:
        """
        Calculate quality score for a component based on pricing structure reliability.
        Higher scores indicate more reliable pricing data.
        """
        try:
            score = 0.0
            
            # Check if component has detailed pricing structure
            if all(key in component for key in ['Kostpris_EP', 'Materialer', 'Timer', 'Takst', 'Påslag_MAT', 'Påslag_UE']):
                score += 0.4  # Has all required fields
                
                # Check if pricing calculations make sense
                calculated_kostpris = component.get('Materialer', 0) + (component.get('Timer', 0) * component.get('Takst', 0))
                if abs(calculated_kostpris - component.get('Kostpris_EP', 0)) < 1.0:  # Allow small rounding differences
                    score += 0.3  # Pricing math is consistent
                
                # Check if påslag values are realistic (typically 0-200%)
                påslag_mat = component.get('Påslag_MAT', 0)
                påslag_ue = component.get('Påslag_UE', 0)
                if 0 <= påslag_mat <= 200 and 0 <= påslag_ue <= 200:
                    score += 0.2  # Påslag values are realistic
                
                # Check if component has source file information
                if component.get('source_file') and component.get('source_file') != 'AI-genereret':
                    score += 0.1  # Has source attribution
                
            else:
                # Component is missing detailed pricing structure
                score = 0.3  # Basic quality for transformed data
                
            return min(score, 1.0)  # Cap at 1.0
            
        except Exception:
            return 0.3  # Default to low quality if calculation fails

    def _initialize_search_system(self):
        """Initialize the search system with the correct knowledge base path"""
        try:
            # Get the path to the trusted knowledge base
            kb_path = os.path.join(os.path.dirname(__file__), '..', 'knowledge_base', 'unified_knowledge_base_backup_before_bnord_fix.json')
            
            # Create a custom search system that knows the correct path
            from component_embeddings import ComponentEmbeddings
            
            # Create embeddings manager with correct path and force regeneration
            embeddings_manager = ComponentEmbeddings(knowledge_base_path=kb_path)
            
            # Force regeneration of embeddings to ensure data integrity
            print("🔄 Forcing regeneration of embeddings to ensure data integrity...")
            success = embeddings_manager.generate_embeddings(force_regenerate=True)
            if success:
                print("✅ Embeddings regenerated successfully with corrected data")
            else:
                print("⚠️ Warning: Failed to regenerate embeddings, using cached data")
            
            # Create semantic search with the embeddings manager
            search_system = SemanticSearch(embeddings_manager=embeddings_manager)
            
            print(f"✓ Search system initialized with knowledge base: {kb_path}")
            return search_system
            
        except Exception as e:
            print(f"⚠️ Warning: Could not initialize search system: {e}")
            print("Search functionality will be limited")
            return None

def main():
    """Main function to run the enhanced AI agent"""
    # Check if OpenAI API key is available
    if not os.getenv("OPENAI_API_KEY"):
        print("❌ Fejl: OPENAI_API_KEY miljøvariabel er ikke sat.")
        print("Sæt venligst din OpenAI API nøgle før du kører denne agent.")
        return
    
    try:
        agent = EnhancedAIAgent()
        agent.run_conversation()
    except Exception as e:
        print(f"❌ Fejl under kørsel af agent: {e}")
        print("Sørg for at du har installeret alle nødvendige pakker: pip install -r requirements.txt")

if __name__ == "__main__":
    main()
