import os
import json
from typing import Dict, List, Any
from dataclasses import dataclass
from datetime import datetime

# LangChain imports
from langchain.agents import initialize_agent, AgentType, Tool
from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.memory import ConversationBufferMemory
from langchain.chains import LLMChain
from langchain.schema import HumanMessage, AIMessage

@dataclass
class ProjectData:
    """Data structure to hold project information"""
    description: str = ""
    square_meters: float = 0.0
    selected_categories: List[str] = None
    category_tasks: Dict[str, List[str]] = None
    
    def __post_init__(self):
        if self.selected_categories is None:
            self.selected_categories = []
        if self.category_tasks is None:
            self.category_tasks = {}

class AIAgent:
    """AI-powered agent using LangChain and OpenAI for construction project planning"""
    
    def __init__(self):
        self.project_data = ProjectData()
        self.categories = self.load_categories()
        self.memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
        
        # Initialize OpenAI LLM
        self.llm = ChatOpenAI(
            temperature=0.1,  # Low temperature for consistent responses
            model="gpt-3.5-turbo"
        )
        
        # Define tools
        self.tools = [
            Tool(
                name="SaveToExcel",
                func=self.save_to_excel,
                description="Saves the final project data to an Excel file"
            ),
            Tool(
                name="GetCategories",
                func=self.get_categories,
                description="Returns the list of available construction categories"
            )
        ]
        
        # Initialize the agent
        self.agent = initialize_agent(
            tools=self.tools,
            llm=self.llm,
            agent=AgentType.CONVERSATIONAL_REACT_DESCRIPTION,
            verbose=True,
            memory=self.memory,
            handle_parsing_errors=True
        )
        
        # Create specialized prompts for different stages
        self.create_prompts()
    
    def load_categories(self) -> List[str]:
        """Load categories from JSON file"""
        try:
            with open("categories.json", "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("categories", [])
        except FileNotFoundError:
            print("Warning: categories.json not found, using default categories")
            return ["nedrivning", "vvs", "elektrisk", "gulv", "vægge", "loft"]
    
    def create_prompts(self):
        """Create specialized prompts for different conversation stages"""
        
        # Initial project description prompt
        self.initial_prompt = PromptTemplate(
            input_variables=["chat_history"],
            template="""Du er en ekspert i byggeprojekter og badeværelsesrenovering. Din opgave er at hjælpe brugeren med at planlægge deres projekt.

{chat_history}

Du skal starte med at spørge om projektets beskrivelse. Stil spørgsmål på dansk og vær venlig og professionel.

Hvis brugeren allerede har beskrevet deres projekt, så spørg om yderligere detaljer som kvadratmeter."""
        )
        
        # Follow-up questions prompt
        self.followup_prompt = PromptTemplate(
            input_variables=["chat_history", "project_description"],
            template="""Baseret på projektbeskrivelsen: "{project_description}"

{chat_history}

Stil relevante opfølgende spørgsmål på dansk for at få mere information om projektet. 
Spørg specifikt om kvadratmeter og andre vigtige detaljer."""
        )
        
        # Category selection prompt
        self.category_prompt = PromptTemplate(
            input_variables=["chat_history", "categories"],
            template="""Her er de tilgængelige byggekategorier:

{categories}

{chat_history}

Præsenter disse kategorier for brugeren på dansk og hjælp dem med at vælge hvilke der er relevante for deres projekt.
Forklar kort hvad hver kategori indebærer."""
        )
        
        # Task gathering prompt
        self.task_prompt = PromptTemplate(
            input_variables=["chat_history", "category"],
            template="""Nu arbejder vi med kategorien: {category}

{chat_history}

Spørg brugeren om hvilke specifikke underopgaver der er nødvendige i denne kategori.
Giv eksempler på typiske opgaver for {category}.
Husk at spørge om der er flere opgaver før du går videre til næste kategori."""
        )
    
    def get_categories(self) -> str:
        """Tool function to get available categories"""
        return f"Tilgængelige kategorier: {', '.join(self.categories)}"
    
    def save_to_excel(self, project_summary: str) -> str:
        """Tool function to save project data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            # Create Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Projekt Budget"
            
            # Add headers
            headers = ["Kategori", "Opgave", "Beskrivelse"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add project info
            ws.cell(row=2, column=1, value="PROJEKT INFO")
            ws.cell(row=3, column=1, value="Beskrivelse")
            ws.cell(row=3, column=2, value=self.project_data.description)
            ws.cell(row=4, column=1, value="Kvadratmeter")
            ws.cell(row=4, column=2, value=self.project_data.square_meters)
            
            # Add tasks
            current_row = 6
            for category, tasks in self.project_data.category_tasks.items():
                for task in tasks:
                    ws.cell(row=current_row, column=1, value=category)
                    ws.cell(row=current_row, column=2, value=task)
                    ws.cell(row=current_row, column=3, value=f"Opgave i {category}")
                    current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ai_projekt_budget_{timestamp}.xlsx"
            wb.save(filename)
            
            return f"Excel-fil gemt som: {filename}"
            
        except Exception as e:
            return f"Fejl ved oprettelse af Excel-fil: {e}"
    
    def run_conversation(self):
        """Main conversation loop using AI agent"""
        print("🤖 Hej! Jeg er din AI-assistent til at planlægge dit byggeprojekt.")
        print("Jeg vil stille dig spørgsmål på dansk og hjælpe dig med at strukturere dit projekt.")
        print("=" * 70)
        
        # Step 1: Get project description
        print("\n📝 **Trin 1: Projektbeskrivelse**")
        user_input = input("Beskriv dit projekt (f.eks. 'Total renovation af badeværelse'): ")
        self.project_data.description = user_input
        
        # Use AI to ask follow-up questions
        response = self.agent.run(f"Brugeren har beskrevet deres projekt som: {user_input}. Stil relevante opfølgende spørgsmål på dansk.")
        print(f"\n🤖 {response}")
        
        # Get square meters
        square_meters_input = input("Hvor mange kvadratmeter er der? (f.eks. 8.5): ")
        try:
            self.project_data.square_meters = float(square_meters_input)
        except ValueError:
            self.project_data.square_meters = 0.0
        
        # Step 2: Present categories with AI assistance
        print(f"\n🏗️ **Trin 2: Kategorier**")
        categories_text = "\n".join([f"{i+1}. {cat}" for i, cat in enumerate(self.categories)])
        
        response = self.agent.run(f"Præsenter disse kategorier for brugeren og forklar hvilke der kunne være relevante for et projekt om: {self.project_data.description}")
        print(f"\n🤖 {response}")
        
        # User selects categories
        print(f"\n{categories_text}")
        selection = input("Indtast numrene på de relevante kategorier (adskilt med komma, f.eks. 1,3,5): ")
        
        try:
            selected_indices = [int(x.strip()) - 1 for x in selection.split(",")]
            self.project_data.selected_categories = [self.categories[idx] for idx in selected_indices if 0 <= idx < len(self.categories)]
        except ValueError:
            self.project_data.selected_categories = []
        
        # Step 3: AI-guided task gathering
        print(f"\n🔨 **Trin 3: Opgaver for hver kategori**")
        
        for category in self.project_data.selected_categories:
            print(f"\n📋 **Kategori: {category}**")
            
            # Use AI to ask about tasks
            response = self.agent.run(f"Spørg brugeren om hvilke specifikke opgaver der er nødvendige i kategorien '{category}' for et projekt om {self.project_data.description}. Giv eksempler på typiske opgaver.")
            print(f"\n🤖 {response}")
            
            tasks = []
            while True:
                task = input(f"Hvilke underopgaver er nødvendige i {category}? ")
                if task.strip():
                    tasks.append(task.strip())
                
                more = input("Er der andet du vil tilføje til denne kategori? (ja/nej): ").lower().strip()
                if more not in ['ja', 'j', 'yes', 'y']:
                    break
            
            self.project_data.category_tasks[category] = tasks
            print(f"✅ {len(tasks)} opgaver tilføjet til {category}")
        
        # Step 4: AI generates project summary and saves to Excel
        print(f"\n💾 **Trin 4: Gem til Excel**")
        
        # Create project summary
        summary = f"""
Projekt: {self.project_data.description}
Kvadratmeter: {self.project_data.square_meters} m²
Kategorier: {', '.join(self.project_data.selected_categories)}
Total opgaver: {sum(len(tasks) for tasks in self.project_data.category_tasks.values())}
        """.strip()
        
        response = self.agent.run(f"Opret et sammendrag af projektet og brug SaveToExcel værktøjet til at gemme dataene. Projekt sammendrag: {summary}")
        print(f"\n🤖 {response}")
        
        print(f"\n✅ Projektet er nu færdigt! Excel-filen er oprettet.")

def main():
    """Main function to run the AI agent"""
    # Check if OpenAI API key is available
    if not os.getenv("OPENAI_API_KEY"):
        print("❌ Fejl: OPENAI_API_KEY miljøvariabel er ikke sat.")
        print("Sæt venligst din OpenAI API nøgle før du kører denne agent.")
        return
    
    try:
        agent = AIAgent()
        agent.run_conversation()
    except Exception as e:
        print(f"❌ Fejl under kørsel af agent: {e}")
        print("Sørg for at du har installeret alle nødvendige pakker: pip install -r requirements.txt")

if __name__ == "__main__":
    main()
